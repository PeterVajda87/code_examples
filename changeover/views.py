from django.shortcuts import render
from django.http import JsonResponse
from django.http import HttpResponse
from django.core.serializers.json import DjangoJSONEncoder
from snowflake_connector.views import snowflake_query
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font
from thingworx.models import MachinesConverter

# Create your views here.
@csrf_exempt
def changeover_view(request):
    fla_arr = MachinesConverter.objects.values_list('short_name').exclude(short_name__exact='')
    data_for_selects = [tup[0] for tup in fla_arr]
    #test = snowflake_query("SELECT OrderNumber FROM SMARTKPI WHERE MACHINE = 'KBLIBFP09-Lanico2PromoticStationThing' AND SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND MACHINE IS NOT NULL")
    #print("DATA")
    #print(test)
    if request.method == "POST":
        request_data = json.loads(request.body)
        request_machine = request_data.get('machine_sel') # stejne pro machine
        
        ####################
        selected_machine_message = list(MachinesConverter.objects.filter(short_name=request_machine).values_list('machine_message_data_name',flat=True)) 
        selected_machine_smartkpi = list(MachinesConverter.objects.filter(short_name=request_machine).values_list('smartkpi_name',flat=True)) 
        
        #print(selected_machine_message,selected_machine_smartkpi)
        selected_machine_message = ','.join(selected_machine_message)
        selected_machine_smartkpi = ','.join(selected_machine_smartkpi)

        selected_machine = selected_machine_message + "', '" + selected_machine_smartkpi
        #####################

        date_from = datetime.strptime(request_data.get('date_from'),'%Y-%m-%d')
        date_to = datetime.strptime(request_data.get('date_to'),'%Y-%m-%d')
        check_buttons = request_data.get('checkArray') #array of shifts (Nocni, Ranni, Odpoledni)

        weeks = True if date_to - date_from >= timedelta(days=14) else False
        data = get_data(weeks, selected_machine, date_from, date_to, check_buttons)
        graph_data = prepare_data_graph(weeks, data, date_from, date_to)
        #print(graph_data)


        request.session['data'] = json.dumps(data[1], cls=DjangoJSONEncoder)
        request.session['selected_machine'] = request_machine.replace(" ", "_") 
        for key, value in request.session.items():
            print('{} => {}'.format(key, value))

        return JsonResponse({
            'data':data,
            'graph_data':graph_data,
        })
    
    if request.method == "GET":
        date_today = datetime.now()

        context = {
            'data_for_selects': data_for_selects,
            'date_today': date_today,
        }
        
        return render(request, "changeover/index.html", context)

def get_data(weeks, selected_machine, date_from, date_to, check_buttons):
    if weeks == False:
        #get_part_shifts = snowflake_query(f"WITH shifts_cte AS (SELECT OrderNumber, PRODUCTIONTIME, LAG(OrderNumber, 1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumber, LAG(PRODUCTIONTIME,1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumberProductionTime, CASE WHEN HOUR(PRODUCTIONTIME) > 22 THEN DATEADD(day,1,PRODUCTIONTIME) ELSE PRODUCTIONTIME END AS shift_day, CASE WHEN HOUR(PRODUCTIONTIME) >= 6 AND HOUR(PRODUCTIONTIME) < 14 THEN 'R' WHEN HOUR(PRODUCTIONTIME) >= 14 AND HOUR(PRODUCTIONTIME) < 22 THEN 'O' ELSE 'N' END AS shift FROM SMARTKPI WHERE OrderNumber IS NOT NULL AND SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND MACHINE = '{sel_machine}' AND PRODUCTIONTIME BETWEEN DATE('{date_from}') AND DATE('{date_to}')) SELECT DATE(shift_day), CAST(AVG(DATEDIFF(second, PreviousOrderNumberProductionTime, ProductionTime)) as integer) FROM shifts_cte WHERE OrderNumber <> PreviousOrderNumber AND shift IN {check_buttons} GROUP BY DATE(shift_day) ORDER BY DATE(shift_day)")
        
        
        #get_part_shifts = snowflake_query(f"WITH shifts_cte AS (SELECT OrderNumber, PRODUCTIONTIME, LAG(OrderNumber, 1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumber, LAG(PRODUCTIONTIME,1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumberProductionTime, CASE WHEN HOUR(PRODUCTIONTIME) > 22 THEN DATEADD(day,1,PRODUCTIONTIME) ELSE PRODUCTIONTIME END AS shift_day, CASE WHEN HOUR(PRODUCTIONTIME) >= 6 AND HOUR(PRODUCTIONTIME) < 14 THEN 'R' WHEN HOUR(PRODUCTIONTIME) >= 14 AND HOUR(PRODUCTIONTIME) < 22 THEN 'O' ELSE 'N' END AS shift FROM SMARTKPI WHERE OrderNumber IS NOT NULL AND SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND MACHINE IN('{selected_machine}') AND PRODUCTIONTIME BETWEEN DATE('{date_from}') AND DATE('{date_to}')), changeover_cte AS ( SELECT MEDIAN(DATEDIFF(minutes, PreviousOrderNumberProductionTime, PRODUCTIONTIME)) AS medianval FROM shifts_cte WHERE PreviousOrderNumber <> OrderNumber ), filtered_changeovers AS ( SELECT * FROM shifts_cte WHERE OrderNumber <> PreviousOrderNumber AND DATEDIFF(minutes, PreviousOrderNumberProductionTime, PRODUCTIONTIME) < (SELECT 2 * AVG(medianval) FROM changeover_cte) ) SELECT DATE(shift_day), CAST(AVG(DATEDIFF(minutes, PreviousOrderNumberProductionTime, ProductionTime)) as integer) FROM filtered_changeovers WHERE OrderNumber <> PreviousOrderNumber AND shift IN ({shifts_to_string(check_buttons)}) GROUP BY DATE(shift_day) ORDER BY DATE(shift_day)")
        get_part_shifts = snowflake_query(f"WITH shifts_cte AS (SELECT OrderNumber, PRODUCTIONTIME, LAG(OrderNumber, 1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumber, LAG(PRODUCTIONTIME,1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumberProductionTime, CASE WHEN HOUR(PRODUCTIONTIME) > 22 THEN DATEADD(day,1,PRODUCTIONTIME) ELSE PRODUCTIONTIME END AS shift_day, CASE WHEN HOUR(PRODUCTIONTIME) >= 6 AND HOUR(PRODUCTIONTIME) < 14 THEN 'R' WHEN HOUR(PRODUCTIONTIME) >= 14 AND HOUR(PRODUCTIONTIME) < 22 THEN 'O' ELSE 'N' END AS shift FROM SMARTKPI WHERE OrderNumber IS NOT NULL AND SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND MACHINE IN('{selected_machine}') AND DATE(PRODUCTIONTIME) <= TO_DATE('{date_to}') AND DATE(PRODUCTIONTIME) >= TO_DATE('{date_from}')), changeover_cte AS ( SELECT MEDIAN(DATEDIFF(minutes, PreviousOrderNumberProductionTime, PRODUCTIONTIME)) AS medianval FROM shifts_cte WHERE PreviousOrderNumber <> OrderNumber ), filtered_changeovers AS ( SELECT * FROM shifts_cte WHERE OrderNumber <> PreviousOrderNumber AND OrderNumber <> '' AND PreviousOrderNumber <> '' AND DATEDIFF(minutes, PreviousOrderNumberProductionTime, PRODUCTIONTIME) < (SELECT 15 * medianval FROM changeover_cte) ) SELECT DATE(shift_day), CAST(AVG(DATEDIFF(minutes, PreviousOrderNumberProductionTime, ProductionTime)) as integer) FROM filtered_changeovers WHERE shift IN ({shifts_to_string(check_buttons)}) GROUP BY DATE(shift_day) ORDER BY DATE(shift_day)")
        get_data_csv = snowflake_query(f"WITH shifts_cte AS (SELECT OrderNumber, PRODUCTIONTIME, LAG(OrderNumber, 1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumber, LAG(PRODUCTIONTIME,1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumberProductionTime, CASE WHEN HOUR(PRODUCTIONTIME) > 22 THEN DATEADD(day,1,PRODUCTIONTIME) ELSE PRODUCTIONTIME END AS shift_day, CASE WHEN HOUR(PRODUCTIONTIME) >= 6 AND HOUR(PRODUCTIONTIME) < 14 THEN 'R' WHEN HOUR(PRODUCTIONTIME) >= 14 AND HOUR(PRODUCTIONTIME) < 22 THEN 'O' ELSE 'N' END AS shift FROM SMARTKPI WHERE OrderNumber IS NOT NULL AND SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND MACHINE IN('{selected_machine}')AND DATE(PRODUCTIONTIME) <= TO_DATE('{date_to}') AND DATE(PRODUCTIONTIME) >= TO_DATE('{date_from}')) SELECT PreviousOrderNumber, OrderNumber, PreviousOrderNumberProductionTime, ProductionTime, DATE(shift_day), shift FROM shifts_cte WHERE OrderNumber <> PreviousOrderNumber AND shift IN ({shifts_to_string(check_buttons)}) ORDER BY DATE(shift_day)")
    else:
        first_week = int(date_from.strftime(" %V"))
        last_week = int(date_to.strftime("%V"))
        #get_part_shifts = snowflake_query(f"WITH shifts_week_cte AS (SELECT OrderNumber, PRODUCTIONTIME, DATEDIFF(second, LAG(PRODUCTIONTIME,1) OVER (ORDER BY PRODUCTIONTIME),PRODUCTIONTIME) AS ProductionDuration, CASE WHEN HOUR(PRODUCTIONTIME) > 22 OR HOUR(PRODUCTIONTIME) < 6 THEN 'N' WHEN HOUR(PRODUCTIONTIME) > 14 AND HOUR(PRODUCTIONTIME) < 22 THEN 'O' ELSE 'R' END AS shift, CASE WHEN HOUR(PRODUCTIONTIME) > 22 THEN DATEADD(day,1,DATE(PRODUCTIONTIME)) ELSE DATE(PRODUCTIONTIME) END AS shift_day FROM SMARTKPI WHERE PRODUCTIONTIME >= \'{datetime.strftime(date_from - timedelta(days=7), '%Y-%m-%d')}\' AND SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND OrderNumber IS NOT NULL AND MACHINE = '{sel_machine}' AND shift IN {check_buttons}) SELECT AVG(ProductionDuration), WEEKISO(DATE(shift_day)) FROM shifts_week_cte WHERE ProductionDuration <= 90 AND WEEKISO(DATE(shift_day)) >= {first_week} AND WEEKISO(DATE(shift_day)) <= {last_week} GROUP BY WEEKISO(DATE(shift_day)) ORDER BY WEEKISO(DATE(shift_day))")
        #ProductionDuration <= 90 ?
        #PRODUCTIONTIME >= \'{datetime.strftime(date_from - timedelta(days=7), '%Y-%m-%d')}\' => ????,
        get_part_shifts = snowflake_query(f"WITH shifts_week_cte AS (SELECT OrderNumber, PRODUCTIONTIME, DATEDIFF(minutes, LAG(PRODUCTIONTIME,1) OVER (ORDER BY PRODUCTIONTIME),PRODUCTIONTIME) AS ProductionDuration, CASE WHEN HOUR(PRODUCTIONTIME) > 22 OR HOUR(PRODUCTIONTIME) < 6 THEN 'N' WHEN HOUR(PRODUCTIONTIME) > 14 AND HOUR(PRODUCTIONTIME) < 22 THEN 'O' ELSE 'R' END AS shift, CASE WHEN HOUR(PRODUCTIONTIME) > 22 THEN DATEADD(day,1,DATE(PRODUCTIONTIME)) ELSE DATE(PRODUCTIONTIME) END AS shift_day FROM SMARTKPI WHERE PRODUCTIONTIME >= \'{datetime.strftime(date_from - timedelta(days=7), '%Y-%m-%d')}\' AND SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND OrderNumber IS NOT NULL AND MACHINE IN('{selected_machine}') AND shift IN ({shifts_to_string(check_buttons)})) SELECT CAST(AVG(ProductionDuration) as integer), WEEKISO(DATE(shift_day)) FROM shifts_week_cte WHERE WEEKISO(DATE(shift_day)) >= {first_week} AND WEEKISO(DATE(shift_day)) <= {last_week} GROUP BY WEEKISO(DATE(shift_day)) ORDER BY WEEKISO(DATE(shift_day))")
        get_data_csv = snowflake_query(f"WITH shifts_week_cte AS (SELECT OrderNumber, PRODUCTIONTIME, DATEDIFF(minutes, LAG(PRODUCTIONTIME,1) OVER (ORDER BY PRODUCTIONTIME),PRODUCTIONTIME) AS ProductionDuration, LAG(OrderNumber, 1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumber, LAG(PRODUCTIONTIME,1) OVER (ORDER BY PRODUCTIONTIME) AS PreviousOrderNumberProductionTime, CASE WHEN HOUR(PRODUCTIONTIME) > 22 OR HOUR(PRODUCTIONTIME) < 6 THEN 'N' WHEN HOUR(PRODUCTIONTIME) > 14 AND HOUR(PRODUCTIONTIME) < 22 THEN 'O' ELSE 'R' END AS shift, CASE WHEN HOUR(PRODUCTIONTIME) > 22 THEN DATEADD(day,1,DATE(PRODUCTIONTIME)) ELSE DATE(PRODUCTIONTIME) END AS shift_day FROM SMARTKPI WHERE PRODUCTIONTIME >= \'{datetime.strftime(date_from - timedelta(days=7), '%Y-%m-%d')}\' AND SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND OrderNumber IS NOT NULL AND MACHINE IN('{selected_machine}')) SELECT PreviousOrderNumber, OrderNumber, PreviousOrderNumberProductionTime, ProductionTime, DATE(shift_day), shift FROM shifts_week_cte WHERE PRODUCTIONTIME BETWEEN DATE('{date_from}') AND DATE('{date_to}') AND OrderNumber <> PreviousOrderNumber AND shift IN ({shifts_to_string(check_buttons)}) ORDER BY DATE(shift_day)")
    return get_part_shifts, get_data_csv

def prepare_data_graph(weeks, data, start_time, end_time):
    if weeks == False:
        end_time = end_time + timedelta(days=1)
        line_x = [start_time.date()+timedelta(days=x) for x in range((end_time - start_time).days)]
        return_list = []
        result_temp = {tick: value for (tick,value) in data[0]}
        for tick in line_x:
            return_list.append(result_temp.get(tick, 0))
    else:
        first_week = int(start_time.strftime(" %V"))
        last_week = int(end_time.strftime("%V"))
        line_x = list(range(first_week, last_week))
        return_list = []
        result_temp = {tick: value for (value,tick) in data[0]}
        for tick in line_x:
            return_list.append(result_temp.get(tick, 0))
    return return_list, line_x
    
def changeover_export(request):
    data = request.session['data']
    name = request.session['selected_machine']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    
    response['Content-Disposition'] = (f"attachment; filename=CHANGEOVER_{name}.xlsx")

    workbook = Workbook()

    worksheet = workbook.active
    
    font = Font(bold=True)
    # pyxl neumoznuje aplikovat zmenu do range
    worksheet['A1'].font = font
    worksheet['B1'].font = font
    worksheet['C1'].font = font
    worksheet['D1'].font = font
    worksheet['E1'].font = font
    worksheet['F1'].font = font

    # Define the titles for columns
    columns = [
        'Previous Part Number',
        'Part Number',
        'Previous Production Time',
        'Production Time',
        'Shift Day',
        'Shift'
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for row in json.loads(data):
        row_num += 1
        #if wrong date in production time
        try:
            pre_prod_time = datetime.strptime(row[2],"%Y-%m-%dT%H:%M:%S.%fZ")
        except:
            pre_prod_time = datetime.strptime(row[2],"%Y-%m-%dT%H:%M:%SZ")
        try:
            curr_prod_time = datetime.strptime(row[3],"%Y-%m-%dT%H:%M:%S.%fZ")
        except:
            curr_prod_time = datetime.strptime(row[3],"%Y-%m-%dT%H:%M:%SZ")
        
        
        row_to_excel = [
            row[0],
            row[1],
            pre_prod_time,
            curr_prod_time,
            row[4],
            row[5]
        ]

        for col_num, cell_value in enumerate(row_to_excel, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


def shifts_to_string(shifts):
    shifts_string = [f"'{shift}'" for shift in shifts] 
    return ", ".join(shifts_string)

