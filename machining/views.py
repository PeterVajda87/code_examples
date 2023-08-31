from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.http import JsonResponse
from thingworx.models import MachinesConverter
from .models import MachiningRecords, Downtime, Machine, DowntimeCategory, RecordedDowntime
import datetime
from django.http import HttpResponse
from django.db.models import F, ExpressionWrapper, fields

from openpyxl import Workbook
from openpyxl.styles import Font
from django.db import connections

@csrf_exempt
def machining_view(request):
    fla_arr = MachinesConverter.objects.values_list('short_name').exclude(short_name__exact='')
    data_for_selects = [tup[0] for tup in fla_arr]
    date_from = datetime.datetime.strptime('02-01-2023', '%d-%m-%Y')
    date_to = datetime.datetime.strptime('15-02-2023',  '%d-%m-%Y')
    #print(data_for_selects)    
    csv_export = list(MachiningRecords.objects.values_list('stroj', 'prostoj', 'duvod', 'zacatek_prostoje','konec_prostoje').exclude(zacatek_prostoje__lte=date_from).exclude(konec_prostoje__gte=date_to))
    print(csv_export)
    
    request.session['date_from'] = date_from.strftime("%d/%m/%Y")
    request.session['date_to'] = date_to.strftime("%d/%m/%Y")

    request.session['machining_data'] = json.dumps(csv_export, cls=DjangoJSONEncoder)
    for key, value in request.session.items():
            print('{} => {}'.format(key, value))
    
    
    if request.method == "POST":
        
        current_date = datetime.date.today()
        request_data = json.loads(request.body)
        stroj = request_data.get("stroj")
        duvod = request_data.get("duvod")
        prostoj = request_data.get("prostoj")
        start_time = datetime.datetime.combine(current_date, datetime.datetime.strptime(request_data.get("time-start"), '%H:%M').time())
        print(start_time)
        end_time = datetime.datetime.combine(current_date, datetime.datetime.strptime(request_data.get("time-end"), '%H:%M').time())
        print(end_time)

        #print(file_name)

        print(stroj,duvod,prostoj,end_time,start_time)

        db_records = MachiningRecords(prostoj=prostoj, stroj=stroj, duvod=duvod, zacatek_prostoje=start_time, konec_prostoje=end_time)
        db_records.save()

        context = {

        }
        return JsonResponse(context)
    if request.method == "GET":
        prostoje = {"Aktivity 5S": ["Údržba 1.stupeň TPM", "Úklid/5s"], "Chybějící materiál": ["Chybějící materiál"], "Chybějící spotřební materiál": ["Chybějící program", "Měřidla", "Nářadí", "Pracovní návodka"], "Ostatní výkon": ["COVID test", "Chybějící operátor", "Operátor v zaučení", "Ověřování v DUMMY kusů", "Prebírání u stroje", "Problém s ThingWorx", "Prodej balení", "Přebírání u stroje", "Předání směny", "Přesun operátorů na jinou linku", "Výroba na jedné paletě"], "Porada/Školení/Trénink": ["Q-First", "Schůzky, školení", "Trénink na pracovišti", "Trénink ostatní"], "Porucha": ["Jiná mechanická porucha (žaluzie atd.)", "Kolize - lidská chyba", "Nejdou nahodit pohony", "Ostatní porucha", "Porucha chlazení, oplachů, čerpadla (emulze)", "Porucha dopravníku špon", "Porucha dveří, zámku", "Porucha laseru", "Porucha monitoru", "Porucha na elektrickém okruhu (včetně čidel, PLC, pohonů, jističů, chráničů atd.)", "Porucha na hydraulickém okruhu", "Porucha osy X, Y, Z, B", "Porucha sondy Renishaw", "Porucha upínací hydrauliky", "Porucha výměny nástrojů", "Porucha výměny palet", "Porucha vřetene", "Prasklá hadice - hydraulika", "Porucha hadice - pneumatika", "Vyšší moc", "Výpadek elektřiny"], "Přestávky/Oběd": ["Přestávka"], "Snížená Dostupnost Zařízení": ["Není obsluha", "Není plánovaná výroba", "Ostatní dostupnost", "Preventivní údržba přípravku", "Preventivní údržba/kontrola stroje", "Výměna nástroje"], "Standardní Přestavba": ["Doseřízení", "KPK", "KPK domezení", "Měření ostatní", "Následná zakázka", "Přeseřízení", "Přestavba na jiný typ", "Seřízení druhé palety"], "Uvolnění/Vzorky/Zkoušky": ["Kvalita", "R&D/CR", "Reklamace", "Technologie", "Výroba vzorků"] }
        db_records = list(MachiningRecords.objects.values('id','duvod','zacatek_prostoje','konec_prostoje', 'prostoj', 'stroj'))

        print(db_records)
        
        context = {
            'data_for_selects':data_for_selects,
            'db_records':db_records,
            'prostoje': prostoje
        }
        return render(request, "machining.html", context)

def machining_export(request):
    machining_data = request.session['machining_data']
    start_time = request.session['date_from']
    end_time = request.session['date_to']

    file_name = start_time + '_' + end_time

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    
    response['Content-Disposition'] = (f"attachment; filename=MACHINING_{file_name}.xlsx")

    workbook = Workbook()

    worksheet = workbook.active
    
    font = Font(bold=True)
    # pyxl neumoznuje aplikovat zmenu do range
    worksheet['A1'].font = font
    worksheet['B1'].font = font
    worksheet['C1'].font = font
    worksheet['D1'].font = font
    worksheet['E1'].font = font

    # Define the titles for columns
    columns = [
        'Machine',
        'Category of Downtime',
        'Cause',
        'Start',
        'End'
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
   
    for row in json.loads(machining_data):
        row_num += 1
        row_to_excel = [
                row[0],
                row[1],
                row[2],
                datetime.datetime.strptime(row[3],"%Y-%m-%dT%H:%M:%S"),
                datetime.datetime.strptime(row[4],"%Y-%m-%dT%H:%M:%S")
            ]

        for col_num, cell_value in enumerate(row_to_excel, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


@csrf_exempt
def downtimes_entry(request, machine):
    if request.method == 'GET':
        machines = list(MachinesConverter.objects.all().values_list('short_name', flat=True).order_by('short_name'))
        
        return render(request, 'machining/downtimes_entry.html', {'machines': machines, 'selected_machine': machine})
    
    
    if request.method == 'POST':
        request_data = json.loads(request.body)
        
        if request_data['subject'] == 'get_categories':
            return JsonResponse({'categories': list(DowntimeCategory.objects.values_list('category', flat=True).order_by('category'))})
        
        if request_data['subject'] == 'get_downtimes':
            return JsonResponse({'downtimes': list(Downtime.objects.filter(category = DowntimeCategory.objects.get(category = request_data['category'])).values_list('downtime', flat=True).order_by('downtime'))})
        
        if request_data['subject'] == 'store_downtime':
            if 'id' in request_data:
                _ = RecordedDowntime.objects.get(id = request_data['id'])
                _.downtime_beginning = request_data['downtimeBeginning']
                _.downtime_end = request_data['downtimeEnd']
                _.downtime = Downtime.objects.get(downtime = request_data['downtime'])
                _.save()
            else:
                _, created = RecordedDowntime.objects.update_or_create(machine = Machine.objects.get_or_create(machine = request_data['machine'])[0], downtime_beginning = request_data['downtimeBeginning'], defaults = {'downtime_end': request_data['downtimeEnd'], 'downtime': Downtime.objects.get(downtime = request_data['downtime'])})
                
            return JsonResponse({'id': _.id})
        
        if request_data['subject'] == 'add_comment':
            pass
        
        if request_data['subject'] == 'delete_downtime':
            RecordedDowntime.objects.filter(id = request_data['id']).delete()
            return JsonResponse({'result': 'ok'})
        
        if request_data['subject'] == 'get_past_records':
            downtimes_list = []
            downtimes = RecordedDowntime.objects.filter(machine = Machine.objects.get(machine = request_data['machine'])).order_by('downtime_end')[:100]
            
            for downtime in downtimes:
                downtime_dict = {}
                downtime_dict['id'] = downtime.id
                downtime_dict['downtimeBeginning'] = downtime.downtime_beginning
                downtime_dict['downtimeEnd'] = downtime.downtime_end
                downtime_dict['downtimeCategory'] = downtime.downtime.category.category
                downtime_dict['downtime'] = downtime.downtime.downtime
                downtimes_list.append(downtime_dict)
                
            return JsonResponse({'downtimes': downtimes_list})    


@csrf_exempt
def report(request, date_from, date_to):
   
    data = RecordedDowntime.objects.filter(downtime_beginning__gte=date_from, downtime_end__lte=date_to).annotate(duration=ExpressionWrapper(F('downtime_end') - F('downtime_beginning'), output_field=fields.DurationField())).order_by('-downtime_beginning')
    
        
    return render(request, 'machining/report.html', {'data': data})
    
    
@csrf_exempt
def visualizations(request):
    
    if request.method == "GET":
        return render(request, 'machining/visualizations.html', {})
    
    if request.method == "POST":
        
        raw_query = f"""
        SELECT 'All', NULL, 0, 0
        
        UNION ALL
        
        SELECT DISTINCT(category), 'All', 0, 0 FROM machining_downtimecategory
        
        UNION ALL
        
        SELECT machining_downtime.downtime, machining_downtimecategory.category, EXTRACT(EPOCH FROM SUM(downtime_end - downtime_beginning))/60, EXTRACT(EPOCH FROM SUM(downtime_end - downtime_beginning))/60
        FROM machining_recordeddowntime
        LEFT JOIN machining_downtime ON machining_recordeddowntime.downtime_id = machining_downtime.id
        LEFT JOIN machining_downtimecategory ON machining_downtime.category_id = machining_downtimecategory.id
        GROUP BY machining_downtime.downtime, machining_downtimecategory.category
        
        UNION ALL

        SELECT CONCAT(machining_machine.machine || ' (' || machining_downtime.downtime) || ') ', machining_downtime.downtime, EXTRACT(EPOCH FROM SUM(downtime_end - downtime_beginning))/60, EXTRACT(EPOCH FROM SUM(downtime_end - downtime_beginning))/60
        FROM machining_recordeddowntime
        LEFT JOIN machining_machine ON machining_recordeddowntime.machine_id = machining_machine.id
        LEFT JOIN machining_downtime ON machining_recordeddowntime.downtime_id = machining_downtime.id
        GROUP BY machining_machine.machine, machining_downtime.downtime
        """

        
        with connections['default'].cursor() as cursor:
            cursor.execute(raw_query)
            
            tuple_results = cursor.fetchall()
            
        list_results = []
            
        for result in tuple_results:
            list_results.append(list(result))
            
        return JsonResponse({'data': list_results})