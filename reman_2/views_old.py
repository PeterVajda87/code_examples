from re import S
from django.db.models import F
from django.shortcuts import render
from django.http.response import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from .models import Customer, Header, Material, MaterialGroup, Operation, Component, Losses, RootMaterialComposition, Yfcorer, SortingResults
from django.db import connections
import json
import datetime


class ReconditionMap:
    def __init__(self, period_start, period_end):
        self.map = {}
        self.period_start = period_start
        self.period_end = period_end
        self.chains = {}
        self.components = []
        self.root_materials = RootMaterialComposition.objects.all().values('material__material')
        self.fg_materials = Material.objects.filter(material_number__in=RootMaterialComposition.objects.all().values_list('compound_material__material_number', flat=True))

    def add_order(self, component_object):
        fm = component_object.pegged_requirement #short for finished material, sorry for that
        sm = component_object.material #short for source material, sorry for that
        self.components.append(component_object)

        self.process_chains(sm, fm)

    def process_chains(self, sm, fm):
        if self.is_root_recondition_material(sm):
            chain_start = sm.compound_material.first().material.material_number
            recondition_root = sm.material_number
            if not chain_start in self.chains:
                self.chains[chain_start] = {}
            if not recondition_root in self.chains[chain_start]:
                self.chains[chain_start][recondition_root] = {}

            if self.is_final_material(fm): ### we have start to end
                final_step = fm.material_number
                self.chains[chain_start][recondition_root][final_step] = {}

            else:
                midsteps = {}
                midstep_index = 0
                material_to_check = fm

                while True:
                    if self.is_final_material(material_to_check):
                        midsteps[midstep_index] = material_to_check.material_number
                        break
                    else:
                        if self.get_next_material(material_to_check).material_number in midsteps.values() or material_to_check.is_root_material:
                            break
                        midstep_index += 1
                        midsteps[midstep_index] = material_to_check.material_number
                        material_to_check = self.get_next_material(material_to_check)

                if len(midsteps.keys()) > 0:
                    if not fm.material_number in self.chains[chain_start][recondition_root]:
                        self.chains[chain_start][recondition_root][fm.material_number] = {}
                    self.chains[chain_start][recondition_root][fm.material_number][midsteps[1]] = {}

                if len(midsteps.keys()) > 1:
                    self.chains[chain_start][recondition_root][fm.material_number][midsteps[1]][midsteps[2]] = {}

                if len(midsteps.keys()) > 2:
                    self.chains[chain_start][recondition_root][fm.material_number][midsteps[1]][midsteps[2]][midsteps[3]] = {}
                

    def is_root_recondition_material(self, sm):
        return sm in self.fg_materials and sm.material_number[0:2] != 'CC'

    def is_final_material(self, fm):
        return not Component.objects.filter(material=fm, order__in=Header.objects.filter(release_date__gte=self.period_start, release_date__lte=self.period_end, order_type='YRM2').values_list('order', flat=True)).exists() or fm.material_number[-1] == 'R'

    def get_next_material(self, material):
        return Component.objects.filter(material=material, order__in=Header.objects.filter(release_date__gte=self.period_start, release_date__lte=self.period_end, order_type='YRM2').values_list('order', flat=True)).first().pegged_requirement 

    def process_results(self):
        for cc_material, first_level_materials in self.chains.items():
            for first_level_material, second_level_materials in first_level_materials.items():
                first_level_materials[first_level_material]['order_quantity'] = 0
                first_level_materials[first_level_material]['quantity_delivered'] = 0
                first_level_materials[first_level_material]['result'] = 0
                first_level_materials[first_level_material]['orders'] = []
                material = Material.objects.get(material_number = first_level_material)

                for component in self.components:
                    if component.material == material:
                        order = Header.objects.get(order = component.order)
                        first_level_materials[first_level_material]['orders'].append(order.order)
                        first_level_materials[first_level_material]['order_quantity'] += order.order_quantity
                        first_level_materials[first_level_material]['quantity_delivered'] += order.quantity_delivered
                        first_level_materials[first_level_material]['result'] = first_level_materials[first_level_material]['quantity_delivered'] / first_level_materials[first_level_material]['order_quantity']

                if type(second_level_materials) == 'dict':
                    for second_level_material, third_level_materials in second_level_materials.items(): 
                        if not second_level_material in ['order_quantity', 'result', 'orders', 'quantity_delivered']:
                            first_level_materials[first_level_material][second_level_material]['order_quantity'] = 0
                            first_level_materials[first_level_material][second_level_material]['quantity_delivered'] = 0
                            first_level_materials[first_level_material][second_level_material]['result'] = 0
                            material = Material.objects.get(material_number = second_level_material)

                            for component in self.components:
                                if component.material == material:
                                    order = Header.objects.get(order = component.order)
                                    first_level_materials[first_level_material][second_level_material]['order_quantity'] += order.order_quantity
                                    first_level_materials[first_level_material][second_level_material]['quantity_delivered'] += order.quantity_delivered
                                    first_level_materials[first_level_material][second_level_material]['result'] = first_level_materials[first_level_material][second_level_material]['quantity_delivered'] / first_level_materials[first_level_material][second_level_material]['order_quantity']

                        if type(third_level_materials) == 'dict':
                            for third_level_material, fourth_level_materials in third_level_materials.items():
                                if not third_level_material in ['order_quantity', 'result', 'orders', 'quantity_delivered']:

                                    first_level_materials[first_level_material][second_level_material][third_level_material]['order_quantity'] = 0
                                    first_level_materials[first_level_material][second_level_material][third_level_material]['quantity_delivered'] = 0
                                    first_level_materials[first_level_material][second_level_material][third_level_material]['result'] = 0
                                    material = Material.objects.get(material_number = third_level_material)

                                    for component in self.components:
                                        if component.material == material:
                                            order = Header.objects.get(order = component.order)
                                            first_level_materials[first_level_material][second_level_material][third_level_material]['order_quantity'] += order.order_quantity
                                            first_level_materials[first_level_material][second_level_material][third_level_material]['quantity_delivered'] += order.quantity_delivered
                                            first_level_materials[first_level_material][second_level_material][third_level_material]['result'] = first_level_materials[first_level_material][second_level_material][third_level_material]['quantity_delivered'] / first_level_materials[first_level_material][second_level_material][third_level_material]['order_quantity']

    def to_json(self):
        return json.dumps(self.chains)



class DisassemblyMap:
    def __init__(self, period_start, period_end):
        self.map = {}
        self.period_start = period_start
        self.period_end = period_end


    def add_order(self, component_object):
        order = Header.objects.get(order=component_object.order)
        rm = component_object.pegged_requirement #short for root material, sorry for that
        cm = component_object.material #short for component material, sorry for that again
        cm_quantity = RootMaterialComposition.objects.get(material=rm, compound_material=cm).compound_material_quantity
        self.process_customer(rm)
        self.process_rm_material_group(rm, order, component_object, cm_quantity)
        self.process_root_material(rm, order, component_object, cm_quantity)
        self.process_cm_material_group(rm, cm, order, component_object, cm_quantity)
        self.process_component_material(rm, cm, order, component_object, cm_quantity)


    def process_customer(self, rm):
        if not rm.customer.customer_name:
            customer_name = "Undefined"
        else:
            customer_name = rm.customer.customer_name

        if not customer_name in self.map:
            self.map[customer_name] = {}
            self.map[customer_name]['table_rows'] = 0


    def process_rm_material_group(self, rm, order, component_object, cm_quantity):
        if not rm.customer.customer_name:
            customer_name = "Undefined"
        else:
            customer_name = rm.customer.customer_name

        if not rm.material_group:
            rm_material_group = "Undefined"
        else:
            rm_material_group = rm.material_group.material_group

        if not rm_material_group in self.map[customer_name]:
            self.map[customer_name][rm_material_group] = {}
            self.map[customer_name][rm_material_group]['qty_in'] = order.order_quantity # how many CC were provided
            self.map[customer_name][rm_material_group]['qty_wanted'] = order.order_quantity * cm_quantity # how many FGs were wanted
            self.map[customer_name][rm_material_group]['qty_out'] = -component_object.quantity_withdrawn # how many FGs were extracted
            self.map[customer_name][rm_material_group]['table_rows'] = 0
        else:
            self.map[customer_name][rm_material_group]['qty_in'] += order.order_quantity
            self.map[customer_name][rm_material_group]['qty_wanted'] += order.order_quantity * cm_quantity
            self.map[customer_name][rm_material_group]['qty_out'] += -component_object.quantity_withdrawn
        
        try:
            self.map[customer_name][rm_material_group]['rate_material_group'] = self.map[customer_name][rm_material_group]['qty_out'] / self.map[customer_name][rm_material_group]['qty_wanted']
        except:
            self.map[customer_name][rm_material_group]['rate_material_group'] = 0

    
    def process_root_material(self, rm, order, component_object, cm_quantity):
        if not rm.customer.customer_name:
            customer_name = "Undefined"
        else:
            customer_name = rm.customer.customer_name

        if not rm.material_group:
            rm_material_group = "Undefined"
        else:
            rm_material_group = rm.material_group.material_group

        if not rm.material_number in self.map[customer_name][rm_material_group]:
            self.map[customer_name][rm_material_group][rm.material_number] = {}
            self.map[customer_name][rm_material_group][rm.material_number]['qty_in'] = order.order_quantity
            self.map[customer_name][rm_material_group][rm.material_number]['qty_wanted'] = order.order_quantity * cm_quantity
            self.map[customer_name][rm_material_group][rm.material_number]['qty_out'] = -component_object.quantity_withdrawn
            self.map[customer_name][rm_material_group][rm.material_number]['table_rows'] = 0 
        else:
            self.map[customer_name][rm_material_group][rm.material_number]['qty_in'] += order.order_quantity
            self.map[customer_name][rm_material_group][rm.material_number]['qty_wanted'] += order.order_quantity * cm_quantity
            self.map[customer_name][rm_material_group][rm.material_number]['qty_out'] += -component_object.quantity_withdrawn

        try:
            self.map[customer_name][rm_material_group][rm.material_number]['rate_root_material'] = self.map[customer_name][rm_material_group][rm.material_number]['qty_out'] / self.map[customer_name][rm_material_group][rm.material_number]['qty_wanted']
        except:
            self.map[customer_name][rm_material_group][rm.material_number]['rate_root_material'] = 0


    def process_cm_material_group(self, rm, cm, order, component_object, cm_quantity):
        if not rm.customer.customer_name:
            customer_name = "Undefined"
        else:
            customer_name = rm.customer.customer_name

        if not cm.material_group:
            cm_material_group = "Undefined"
        else:
            cm_material_group = cm.material_group.material_group

        if not rm.material_group:
            rm_material_group = "Undefined"
        else:
            rm_material_group = rm.material_group.material_group

        if not cm_material_group in self.map[customer_name][rm_material_group][rm.material_number]:
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group] = {}
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['no_of_different_cm'] = 0
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['qty_in'] = order.order_quantity
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['qty_wanted'] = order.order_quantity * cm_quantity
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['qty_out'] = -component_object.quantity_withdrawn
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['table_rows'] = 0
        else:
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['qty_in'] += order.order_quantity
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['qty_wanted'] += order.order_quantity * cm_quantity
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['qty_out'] += -component_object.quantity_withdrawn


    def process_component_material(self, rm, cm, order, component_object, cm_quantity):
        if not rm.customer.customer_name:
            customer_name = "Undefined"
        else:
            customer_name = rm.customer.customer_name

        if not cm.material_group:
            cm_material_group = "Undefined"
        else:
            cm_material_group = cm.material_group.material_group

        if not rm.material_group:
            rm_material_group = "Undefined"
        else:
            rm_material_group = rm.material_group.material_group

        if not cm.material_number in self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]:
            self.map[customer_name]['table_rows'] += 1
            self.map[customer_name][rm_material_group]['table_rows'] += 1
            self.map[customer_name][rm_material_group][rm.material_number]['table_rows'] += 1
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['table_rows'] += 1
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['no_of_different_cm'] += 1
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number] = {}
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['qty_in'] = order.order_quantity
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['qty_wanted'] = order.order_quantity * cm_quantity
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['qty_out'] = -component_object.quantity_withdrawn
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['q3'] = {}
        else:
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['qty_in'] += order.order_quantity
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['qty_wanted'] += order.order_quantity * cm_quantity
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['qty_out'] += -component_object.quantity_withdrawn


        if Losses.objects.filter(assembly=cm).exists():
            q3 = Losses.objects.filter(assembly=cm).filter(order__in=Header.objects.filter(release_date__gte=self.period_start, release_date__lte=self.period_end).values_list('order', flat=True))

            for order in q3:
                self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['q3'][order.order] = order.quantity
        else:
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['q3'] = {}
            

        try:
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['rate_component_material'] = self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['qty_out'] / self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['qty_wanted']
        except:
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group][cm.material_number]['rate_component_material'] = 0

        try:
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['rate_cm_group'] = self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['qty_out'] / (self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['qty_wanted'] / self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['no_of_different_cm'])
        except:
            self.map[customer_name][rm_material_group][rm.material_number][cm_material_group]['rate_cm_group'] = 0


    def get_map(self):
        return self.map


    def to_json(self):
        return json.dumps(self.map)


class MaterialMap:
    def __init__(self):
        self.materials = set()

    def add_material(self, remanufactured_material):
        self.materials.add(remanufactured_material)

    def show_materials(self):
        return self.all_materials


class RemanufacturedMaterial:
    def __init__(self, materialObject):
        self.material_object = materialObject
        self.is_root_material = False
        self.composition = {} ### only valid for root materials
        self.check_if_root_material()

    def __eq__(self, other):
        return self.material_object.material_number == other.material_object.material_number

    def __str__(self):
        return self.material_object.material_number

    def __repr__(self):
        return self.material_object.material_number

    def __hash__(self):
        return hash(self.material_object.material_number)

    def check_if_root_material(self):
        if self.material_object.material_number[0:2] == 'CC':
            self.is_root_material = True
            self.material_object.is_root_material = True
            self.material_object.save()
            self.update_composition()

    def update_composition(self):
        components = Component.objects.filter(pegged_requirement = self.material_object).exclude(material = self.material_object)
        for component in components:
            if not RootMaterialComposition.objects.filter(material=self.material_object, compound_material=component.material).exists():
                    RootMaterialComposition.objects.create(
                        material = self.material_object,
                        compound_material = component.material,
                    )
                    self.material_object.refreshed = datetime.datetime.now()
                    self.material_object.save()
                    component.material.refreshed = datetime.datetime.now()
                    component.material.save()

    def get_composition(self):
        return RootMaterialComposition.objects.filter(material=self.material_object)



def index(request):

    headers = Header.objects.filter(release_date__gte='2022-07-01')

    return render(request, 'reman_2/index.html', {'headers': headers})


@csrf_exempt
def root_materials(request):

    if request.method == "GET":

        # materials = Material.objects.all()

        # for material in materials:
        #     remanufactured_material = RemanufacturedMaterial(material)

        root_materials = Material.objects.filter(is_root_material=True).order_by('material_number')
        customers = Customer.objects.all()
        material_groups = MaterialGroup.objects.all()

        context = {
            'root_materials': root_materials,
            'customers': customers,
            'material_groups': material_groups,
        }

        return render(request, 'reman_2/root_materials.html', context)

    if request.method == "POST":
        request_data = json.loads(request.body)
        if 'root-material-composition-id' in request_data:
            RootMaterialComposition.objects.filter(id=request_data['root-material-composition-id']).update(compound_material_quantity=request_data['compound-material-quantity'])
        
        if 'field' in request_data:
            material = Material.objects.get(id=request_data['id'])
            if request_data['field'] == 'material-group':
                material_group = MaterialGroup.objects.get(material_group=request_data['value']) if MaterialGroup.objects.filter(material_group=request_data['value']).exists() else None
                material.material_group = material_group
            if request_data['field'] == 'customer':
                customer = Customer.objects.get(customer_name=request_data['value'])
                material.customer = customer
            material.save()
    

        return JsonResponse({'status': 'okay'})
    

def sorting(request):
    period_start = request.session.get('date-from')
    if not period_start:
        period_start = (datetime.date.today().replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
    period_end = request.session.get('date-to')
    if not period_end:
        period_end = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)
    materials = Material.objects.filter(is_root_material=True)
    material_groups = []
    [material_groups.append(material.material_group) for material in materials if not material.material_group in material_groups and material.material_group]
    customers = Customer.objects.all()

    ### this is for single material (and, logically) customer ###

    for material in materials:
        if Yfcorer.objects.filter(material=material, change_date__gte=period_start, change_date__lte=period_end).exists():
            SortingResults.objects.update_or_create(
                period_start = period_start,
                period_end = period_end,
                material = material,
                customer = material.customer,
                defaults = {'result': Yfcorer.objects.filter(material=material, change_date__gte=period_start, change_date__lte=period_end, core_inventory_flag=1).count() / Yfcorer.objects.filter(material=material, change_date__gte=period_start, change_date__lte=period_end).count() if Yfcorer.objects.filter(material=material, change_date__gte=period_start, change_date__lte=period_end).count() > 0 else 0}
            ) 

    ### this is for material group AND customer ###

    for customer in customers:
        for material_group in material_groups:

            if Yfcorer.objects.filter(material__in=Material.objects.filter(material_group=material_group, customer=customer, is_root_material=True)).exists():

                SortingResults.objects.update_or_create(
                    period_start = period_start,
                    period_end = period_end,
                    material_group = material_group,
                    customer = customer,
                    defaults = {'result': Yfcorer.objects.filter(material__in=Material.objects.filter(material_group=material_group, is_root_material=True, customer=customer), change_date__gte=period_start, change_date__lte=period_end, core_inventory_flag=1).count() / Yfcorer.objects.filter(material__in=Material.objects.filter(material_group=material_group, customer=customer, is_root_material=True), change_date__gte=period_start, change_date__lte=period_end).count() if Yfcorer.objects.filter(material__in=Material.objects.filter(material_group=material_group, customer=customer, is_root_material=True), change_date__gte=period_start, change_date__lte=period_end).count() > 0 else 0}
                )

    context = {
        'results': SortingResults.objects.filter(period_start=period_start, period_end=period_end, material__isnull=False).order_by('customer', 'material__material_group__material_group'),
        'material_groups_results': {f"{customer}{material_group}": v for (customer, material_group, v) in list(SortingResults.objects.filter(period_start=period_start, period_end=period_end, material__isnull=True).order_by('customer__customer_name', 'material_group').values_list('customer__customer_name', 'material_group__material_group', 'result'))}
    }

    return render(request, 'reman_2/sorting.html', context)

@csrf_exempt
def settings(request):
    if request.method == "GET":
        customers = Customer.objects.all().order_by('customer_name')
        material_groups = MaterialGroup.objects.all().order_by('material_group')
        date_from = request.session.get('date-from')
        date_to = request.session.get('date-to')

        context = {
            'customers': customers,
            'material_groups': material_groups,
            'date_from': date_from,
            'date_to': date_to,
        }

        return render(request, 'reman_2/settings.html', context)

    if request.method == "POST":
        request_data = json.loads(request.body)

        if request_data.get('data-type') == 'customer-name':
            Customer.objects.update_or_create(customer_name=request_data.get('data-value'))
        if request_data.get('data-type') == 'material-group':
            MaterialGroup.objects.update_or_create(material_group=request_data.get('data-value'))
        if 'dateFrom' in request_data:
            request.session['date-from'] = request_data.get('dateFrom')
        if 'dateTo' in request_data:
            request.session['date-to'] = request_data.get('dateTo')

        return JsonResponse({'status': 'okay'})
      

@csrf_exempt
def other_materials(request):

    if request.method == 'GET':

        non_root_materials = Material.objects.exclude(is_root_material = True).order_by('material_number')
        material_groups = MaterialGroup.objects.all()

        return_dict = {}

        for material in non_root_materials:
            if material.compound_material.exists():
                if not material in return_dict:
                    return_dict[material] = []
                return_dict[material] = [_.material.material_number for _ in material.compound_material.all()]  

        context = {
            'other_materials': return_dict,
            'material_groups': material_groups,
        }

        return render(request, 'reman_2/other_materials.html', context)

    if request.method == 'POST':

        request_data = json.loads(request.body)
        material = Material.objects.get(id=request_data['id'])
        new_material_group = MaterialGroup.objects.get(material_group=request_data['value']) if MaterialGroup.objects.filter(material_group=request_data['value']).exists() else None
        material.material_group = new_material_group
        material.save()


def disassembly(request):
    period_start = request.session.get('date-from')
    if not period_start:
        period_start = (datetime.date.today().replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
    period_end = request.session.get('date-to')
    if not period_end:
        period_end = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)

    orders_in_period = list(Header.objects.filter(release_date__gte=period_start, release_date__lte=period_end, order_type='YRM4').values_list('order', flat=True))
    components = Component.objects.filter(order__in=orders_in_period).exclude(pegged_requirement=F('material'))

    disassembly_map = DisassemblyMap(period_start, period_end)

    for component in components:
        disassembly_map.add_order(component)

    context = {
        'results': disassembly_map.to_json(),
    }

    return render(request, 'reman_2/disassembly.html', context)


@csrf_exempt
def recondition(request):
    if request.method == "GET":

        period_start = request.session.get('date-from')
        if not period_start:
            period_start = (datetime.date.today().replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
        period_end = request.session.get('date-to')
        if not period_end:
            period_end = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)

        orders_in_period = list(Header.objects.filter(release_date__gte=period_start, release_date__lte=period_end, order_type='YRM2').values_list('order', flat=True))
        components = Component.objects.filter(order__in=orders_in_period).exclude(pegged_requirement=F('material')).exclude(material__material_number__startswith='CC')

        recondition_map = ReconditionMap(period_start, period_end)
        
        for component in components:
            recondition_map.add_order(component)

        recondition_map.process_results()

        context = {
            'results': recondition_map.to_json()
        }
        
        return render(request, 'reman_2/recondition.html', context)


    if request.method == "POST":

        return JsonResponse()



@csrf_exempt
def upload_excel(request):
    if request.method == "POST":
        return_message = ""
        for file in request.FILES:
            wb = load_workbook(filename=request.FILES[file].file)
            ws = wb.active
            addresses = {}
            for row in ws.iter_rows(min_row = 1, max_row = 1):
                for index, cell in enumerate(row):
                    addresses[cell.value] = index

            iter_rows = ws.rows
            next(iter_rows)

            if "headers" in file:
                try:
                    for row in iter_rows:
                        material, created = Material.objects.get_or_create(material_number = row[addresses['Material Number']].value)
                        Header.objects.update_or_create(order = row[addresses['Order']].value, defaults = {
                            'order_type': row[addresses['Order Type']].value, 
                            'material_number': material, 
                            'order_quantity': row[addresses['Order quantity (GMEIN)']].value, 
                            'quantity_delivered': row[addresses['Quantity Delivered (GMEIN)']].value, 
                            'rework_quantity': row[addresses['Rework qty (GMEIN)']].value, 
                            'confirmed_scrap': row[addresses['Confirmed scrap (GMEIN)']].value, 
                            'system_status': row[addresses['System Status']].value, 
                            'release_date': row[addresses['Release date (actual)']].value, 
                            'actual_finish_date': row[addresses['Actual finish date']].value,
                            })
                    return_message += 'Nahrán soubor s hlavičkami. '
                except:
                    return_message += 'Soubor headers neobsahoval potřebné sloupce. '

            if "operations" in file:
                try:
                    for row in iter_rows:
                        material, created = Material.objects.get_or_create(material_number = row[addresses['Material']].value)
                        Operation.objects.update_or_create(order = row[addresses['Order']].value, activity = row[addresses['Activity']].value, defaults = {
                            'activity': row[addresses['Activity']].value, 
                            'material_number': material, 
                            'work_center': row[addresses['Work center']].value,
                            'operation_text': row[addresses['Operation short text']].value,
                            'operation_quantity': row[addresses['Operation Quantity (MEINH)']].value,
                            'confirmed_yield': row[addresses['Confirmed yield (MEINH)']].value,
                            'scrap': row[addresses['Scrap']].value,
                            'rework_quantity': row[addresses['Rework qty (MEINH)']].value,
                            }) 
                    return_message += 'Nahrán soubor s operacemi. '
                except:
                    return_message += 'Soubor operations neobsahoval potřebné sloupce. '

            if "components" in file:
                try:
                    for row in iter_rows:
                        pegged_requirement, created = Material.objects.get_or_create(material_number = row[addresses['Pegged requirement']].value)
                        material, created = Material.objects.get_or_create(material_number = row[addresses['Material']].value)
                        Component.objects.update_or_create(order = row[addresses['Order']].value, material = material, defaults = {
                            'pegged_requirement': pegged_requirement,
                            'requirement_quantity': row[addresses['Requirement Quantity (EINHEIT)']].value,
                            'quantity_withdrawn': row[addresses['Quantity withdrawn (EINHEIT)']].value
                            })
                    return_message += 'Nahrán soubor s komponentami. '
                except:
                    return_message += 'Soubor components neobsahoval potřebné sloupce. '

            if "q3" in file:
                try:
                    for row in iter_rows:
                        material, created = Material.objects.get_or_create(material_number = row[addresses['Material']].value)
                        assembly, created = Material.objects.get_or_create(material_number = row[addresses['Assembly']].value)
                        Losses.objects.update_or_create(order = row[addresses['Production order']].value, material = material, damage = row[addresses['Damage Code']].value, defaults = {
                            'assembly': assembly,
                            'quantity': row[addresses['DefectiveQty (internal)']].value,
                            })
                    return_message += 'Nahrán soubor s Q3. '
                except:
                    return_message += 'Soubor q3 neobsahoval potřebné sloupce. '

            if "yfcorer" in file:
                try:
                    for row in iter_rows:
                        material, created = Material.objects.get_or_create(material_number = row[addresses['Core Category Number']].value)
                        if row[addresses['Core Inventory Flag']].value in [1, 3, "1", "3"]:
                            Yfcorer.objects.update_or_create(core_shipment_id = row[addresses['Core Shipment ID']].value, pallet_number = row[addresses['Pallet Number']].value, core_return_item_number=row[addresses['Core Return Item No.']].value, defaults = {
                                'change_date': row[addresses['Change Date']].value, 
                                'core_inventory_flag': row[addresses['Core Inventory Flag']].value,
                                'material': material
                                }) 
                    return_message += 'Nahrán soubor s yfcorer. '
                except:
                    return_message += 'Soubor yfcorer neobsahoval potřebné sloupce. '


            if return_message == "":
                return JsonResponse({'response': 'Nebyly nalezeny soubory obsahující validní data (headers, operations, components, q3, yfcorer)'})

        materials = Material.objects.all()

        for material in materials:
            remanufactured_material = RemanufacturedMaterial(material)

        return JsonResponse({'response': return_message})

    return render(request, 'reman_2/upload_excel.html', {})


