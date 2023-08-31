from django.shortcuts import render
from .models import QualityNokDescription
from thingworx.models import MachinesConverter
from snowflake_connector.views import snowflake_query
import json
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime, timedelta
from collections import Counter
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

# Create your views here.
@csrf_exempt
def kvalita_views(request):
    #old way
    #fla_arr = snowflake_query("SELECT DISTINCT MACHINE FROM SMARTKPI_MACHINEMESSAGEDATA WHERE SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND MACHINE IS NOT NULL AND CREATIONTIME > '2021-08-09' AND MESSAGETYPE1 = 'Operator Screen'")
    #print(fla_arr)
    #data_for_selects = [tup[0] for tup in fla_arr]
    ##################
    #new way
    fla_arr = MachinesConverter.objects.values_list('short_name').exclude(short_name__exact='')
    data_for_selects = [tup[0] for tup in fla_arr]
    print(data_for_selects)
    ##############################################

    if request.method == "POST":
        request_data = json.loads(request.body)
        request_machine = request_data.get('machine_sel') # stejne pro machine
        #old way
        #selected_machine = "KBLIB" + request_machine + "StationThing"
        #new way
        selected_machine_message = list(MachinesConverter.objects.filter(short_name=request_machine).values_list('machine_message_data_name',flat=True)) 
        selected_machine_smartkpi = list(MachinesConverter.objects.filter(short_name=request_machine).values_list('smartkpi_name',flat=True)) 
        
        print(selected_machine_message,selected_machine_smartkpi)
        selected_machine_message = ','.join(selected_machine_message)
        selected_machine_smartkpi = ','.join(selected_machine_smartkpi)

        #sql formatovani
        selected_machine = selected_machine_message + "', '" + selected_machine_smartkpi
        

        #print(selected_machine_message)
        #print(selected_machine_smartkpi)

        #selected_machine = MachinesConverter.objects.filter(short_name=request_machine).values_list('machine_message_data_name','smartkpi_name')
    
        selected_nok = request_data.get('NOK')

        date_from = request_data.get('date_from')
        date_to = request_data.get('date_to')
        #data_for_graph = snowflake_query(f"SELECT MESSAGE FROM SMARTKPI_MACHINEMESSAGEDATA WHERE SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND MACHINE = '{selected_machine}' AND MACHINE IS NOT NULL AND CREATIONTIME < '{date_to}' AND CREATIONTIME > '{date_from}' AND MESSAGETYPE1 = 'Operator Screen'")
        
        data_for_graph = snowflake_query(f"SELECT MESSAGE FROM SMARTKPI_MACHINEMESSAGEDATA WHERE SOURCESYSTEM = 'smartproductionLIB.corp.knorr-bremse.com' AND MACHINE IN('{selected_machine}') AND MACHINE IS NOT NULL AND CREATIONTIME < '{date_to}' AND CREATIONTIME > '{date_from}' AND MESSAGETYPE1 = 'Operator Screen'")
        print(data_for_graph)

        nok_desc = QualityNokDescription.objects.values_list('nok_button', 'description')
        nok_dict = {button: desc for (button, desc) in nok_desc}

        line_x = [nok_dict[tup[0]] if tup[0] in nok_dict else tup[0] for tup in data_for_graph] #nok desc graph
        line_y = get_graph_values(line_x)

        nok_dates = get_nok_date(selected_machine,selected_nok,nok_dict,date_from,date_to)

        sorted_list = SortSub(line_y)
        pareto_percentage = compute_pareto_percentage(sorted_list)

        machine_noks_export = request_machine + '_' + date_from + '_' + date_to
        try:
            selected_nok_export = selected_nok + '_' + date_from + '_' + date_to
        except:
            selected_nok_export = ''

        print(sorted_list)
        request.session['sel_machine_name'] = machine_noks_export 
        request.session['sel_nok_name'] = selected_nok_export 
        request.session['data'] = sorted_list
        request.session['specific_noks'] = nok_dates
        

        n = 2
        for i in range(len(pareto_percentage)):
            sorted_list[i].insert(n, pareto_percentage[i])
            n += 2

        return JsonResponse({
                'data_for_selects':data_for_selects,
                'line_y':line_y,
                'nok_dates':nok_dates,
        })
    if request.method == "GET":
        date_today = datetime.now()

        context = {
            'data_for_selects': data_for_selects,
            'date_today': date_today,
        }
        return render(request, "kvalita/index.html", context)

def get_graph_values(line_x):
    line_y = []
    count = Counter(line_x)
    for key, value in count.items():
        temp = [key,value]
        line_y.append(temp)
    return line_y

def get_nok_date(selected_machine,selected_nok,nok_dict,date_from,date_to):
    #vyhledavani v db pomoci orig nazvu ulozeneho v dict
    value_list = list(nok_dict.values())
    if(selected_nok in value_list):
        kb_value = [k for k, v in nok_dict.items() if v == selected_nok]
        selected_nok = kb_value[0]
    
    #vyhledavaci query pro nok dates
    #selected_nok_occurrence = snowflake_query(f"SELECT DATE(MESSAGETIME), COUNT(*) FROM SMARTKPI_MACHINEMESSAGEDATA WHERE MACHINE = '{selected_machine}' AND MACHINE IS NOT NULL AND MESSAGETYPE1 = 'Operator Screen' AND MESSAGE = '{selected_nok}' GROUP BY 1 ORDER BY 1 ASC")
    #selected_nok_occurrence = snowflake_query(f"SELECT DATE(MESSAGETIME), COUNT(*) FROM SMARTKPI_MACHINEMESSAGEDATA WHERE MACHINE = '{selected_machine}' AND MACHINE IS NOT NULL AND MESSAGETYPE1 = 'Operator Screen' AND MESSAGE = '{selected_nok}' AND MESSAGETIME BETWEEN '{date_from}' AND '{date_to}' GROUP BY 1 ORDER BY 1 ASC")
    start_time = datetime.strptime(date_from,'%Y-%m-%d').date()
    end_time = datetime.strptime(date_to,'%Y-%m-%d').date()
    nok_occurrence_query = snowflake_query(f"SELECT TO_CHAR(MESSAGETIME,'YYYY-MM-DD'), COUNT(*) FROM SMARTKPI_MACHINEMESSAGEDATA WHERE MACHINE IN('{selected_machine}') AND MACHINE IS NOT NULL AND MESSAGETYPE1 = 'Operator Screen' AND MESSAGE = '{selected_nok}' AND MESSAGETIME BETWEEN '{start_time}' AND '{end_time}' GROUP BY 1 ORDER BY 1 ASC")

    generate_dates = list(str(start_time+timedelta(days=x)) for x in range((end_time-start_time).days))
    return_list = []
    result_temp = {tick: value for (tick,value) in nok_occurrence_query}
    for tick in generate_dates:
        return_list.append(result_temp.get(str(tick), 0))
    return return_list, generate_dates

def SortSub(sub_list):
    sub_list.sort(key = lambda x: x[1], reverse=True)
    return sub_list

def compute_pareto_percentage(sorted_list):
    values_sum = 0
    prom = 0
    localProcento = []

    for i in range(len(sorted_list)):
        values_sum += sorted_list[i][1]
    for j in range(len(sorted_list)):
        prom = prom + sorted_list[j][1]
        localProcento.append(round((prom/values_sum) * 100, 2))
    return localProcento

def noks_export(request):

    data = request.session['data']
    name = request.session['sel_machine_name']
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    
    response['Content-Disposition'] = (f"attachment; filename=KVALITA_{name}.xlsx")

    workbook = Workbook()

    worksheet = workbook.active
    
    font = Font(bold=True)
    # pyxl neumoznuje aplikovat zmenu do range
    worksheet['A1'].font = font
    worksheet['B1'].font = font
    worksheet['C1'].font = font

    # Define the titles for columns
    columns = [
        'NOK',
        'Number of occurences',
        'Pareto Percentage'
        
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for row in data:
        row_num += 1

        row_to_excel = [
            row[0],
            row[1],
            row[2]
        ]

        for col_num, cell_value in enumerate(row_to_excel, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def specific_nok_export(request):
    specific_noks = request.session['specific_noks']
    name = request.session['sel_nok_name']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    
    response['Content-Disposition'] = (f"attachment; filename=KVALITA_{name}.xlsx")

    workbook = Workbook()

    worksheet = workbook.active
    
    font = Font(bold=True)
    # pyxl neumoznuje aplikovat zmenu do range
    worksheet['A1'].font = font
    worksheet['B1'].font = font

    # Define the titles for columns
    columns = [
        'Date',
        'Number of occurences'
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for row in specific_noks:
        row_num += 1
        for row in zip(specific_noks[1], specific_noks[0]):
            worksheet.append(row)

    workbook.save(response)

    return response
