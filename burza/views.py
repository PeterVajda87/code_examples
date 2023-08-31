from django.http.response import JsonResponse
from django.shortcuts import render
from django.http import HttpResponseRedirect, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Worker, Offer, Request, Assignment, PendingRequest, CostCenter, BurzaUser
from django.contrib import messages
import datetime
import time
import calendar
from django.db.models import Q, manager
from django.contrib.auth.decorators import login_required, permission_required
import calendar
import smtplib
from django.contrib.auth.models import User
from django.utils.translation import get_language_from_request
import html
from django.db.models import Count
from django.views.decorators.clickjacking import xframe_options_exempt
from openpyxl import Workbook
from django.db import connection
import json


@csrf_exempt
def edit_workers(request):

    if request.method == "POST":
        if 'csrfmiddlewaretoken' in request.POST:
            personal_number = request.POST.get('personal-number')
            name = request.POST.get('name')
            cost_center = request.POST.get('cost-center')
            manager = Worker.objects.filter(cost_center=cost_center).first().manager
            cost_center_object = CostCenter.objects.get(number=cost_center)

            worker, updated = Worker.objects.update_or_create(personal_number=personal_number, defaults = {'name': name, 'manager': manager, 'cost_center': cost_center, 'costcenter': cost_center_object})
        
        else:
            data = json.loads(request.body)
            for worker_id, values in data.items():
                cost_center_object = CostCenter.objects.get(number=data[worker_id]['cost_center'])
                worker, updated = Worker.objects.update_or_create(personal_number=data[worker_id]['personal_number'], defaults = {'name': data[worker_id]['name'], 'costcenter': cost_center_object, 'left': data[worker_id]['left']})      

        
    operators = Worker.objects.all().order_by('name')

    context = {
        'operators': operators,
    }   

    return render(request, 'edit_workers.html', context)


def export_controlling(request):

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=burza-{date}.xlsx'.format(
        date=datetime.datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'burza_pohyby'

    # Define the titles for columns
    columns = [
        'Id',
        'Osobní číslo',
        'Zdrojové středisko',
        'Cílové středisko',
        'Realizováno',
        'Hodiny',
        'Den',
        'Směna',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    custom_query = "SELECT burza_assignment.id, burza_worker.personal_number AS osobni_cislo, t1.number AS zdrojove_stredisko, t2.number AS cilove_stredisko, burza_assignment.realised AS realizovano, burza_assignment.hours AS hodiny, COALESCE(burza_request.day, burza_offer.day) AS den_vypujcky, COALESCE(burza_request.shift, burza_offer.shift) AS smena FROM burza_assignment LEFT JOIN burza_worker ON burza_worker.id = burza_assignment.worker_id LEFT JOIN burza_costcenter t1 ON burza_worker.costcenter_id = t1.id LEFT JOIN burza_costcenter t2 ON burza_assignment.target_costcenter_id = t2.id LEFT JOIN burza_request ON burza_request.id = burza_assignment.request_id LEFT JOIN burza_offer ON burza_offer.id = burza_assignment.offer_id ORDER BY den_vypujcky DESC"

    with connection.cursor() as cursor:
        cursor.execute(custom_query)
        rows = cursor.fetchall()

    for row in rows:
        row_num += 1

        row_to_excel = [
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[7]
        ]

        for col_num, cell_value in enumerate(row_to_excel, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response

@xframe_options_exempt
def show_calendar(request, selected_month=datetime.date.today().month, selected_year=2023):

    c = calendar.Calendar(firstweekday=0)

    days_with_offers = []

    for day in c.itermonthdays4(year=selected_year, month=selected_month):
        days_with_offers.append(day)

    weeks = c.monthdatescalendar(year=selected_year, month=selected_month)


    cal_to_web = {}
    for week in weeks:
        cal_to_web[weeks.index(week)] = {}
        for day in week:
            cal_to_web[weeks.index(week)][day] = {}
            cal_to_web[weeks.index(week)][day]['Offers'] = []
            cal_to_web[weeks.index(week)][day]['Requests'] = []
            cal_to_web[weeks.index(week)][day]['Current'] = False
            if day.month == selected_month:
                cal_to_web[weeks.index(week)][day]['Current'] = True
            cal_to_web[weeks.index(week)][day]['Offers'].append(Offer.objects.filter(day=day).values('id', 'worker__name', 'worker__costcenter__name', 'worker__personal_number', 'worker__cost_center', 'shift', 'fullfiled', 'start', 'end', 'standard', 'assignment__target_costcenter__name', 'assignment__created_by__last_name', 'created_by__last_name').order_by('fullfiled', '-shift', '-standard'))
            cal_to_web[weeks.index(week)][day]['Requests'].append(Request.objects.filter(day=day).values('target_costcenter__name', 'shift', 'fullfiled', 'id', 'start', 'end', 'standard', 'assignment__worker__name', 'assignment__hours', 'assignment__created_by__last_name', 'created_by__last_name').order_by('fullfiled', '-shift', '-standard')) 
            cal_to_web[weeks.index(week)][day]['Full'] = True

    for week in weeks:
        for day in week:
            for offer in cal_to_web[weeks.index(week)][day]['Offers']:
                for item in offer:
                    if item['fullfiled'] is False:
                        cal_to_web[weeks.index(week)][day]['Full'] = False

    if request.method == 'GET':
        pass

    context = {
        "selected_month": selected_month,
        "selected_year": selected_year,
        "cal_to_web": cal_to_web,
        "today": datetime.date.today(),
    }

    if request.user.is_authenticated:
        context['authenticated'] = True
        if Worker.objects.get(name = request.user.last_name + ' ' + request.user.first_name).is_manager:
            context['manager'] = True

    return render(request, 'show_calendar.html', context)

@login_required
def add_available(request, selected_day=datetime.date.today().day, selected_month=datetime.date.today().month, selected_year=datetime.date.today().year):

    context = {}

    if request.user.is_authenticated:
        context['user_name'] = str(request.user.last_name) + ' ' + (request.user.first_name)
    
    userObject = Worker.objects.get(name = context['user_name'])

    if not userObject.is_manager:
        users_manager = Worker.objects.get(name = userObject.name).manager
        worker_objects = Worker.objects.filter(manager__in=[users_manager, userObject.name], not_available = False, left = False).order_by('name')
        extra_workers = Worker.objects.filter(cost_center = userObject.cost_center).order_by('name')
        worker_objects = worker_objects.union(extra_workers).order_by('name')
    
    if userObject.is_manager:
        worker_objects = Worker.objects.filter(manager = context['user_name'], not_available = False, left = False).order_by('name')

    if userObject.name == 'Szentkirályi Martin':
        worker_objects = Worker.objects.filter(manager = context['user_name'], not_available = False, left = False).order_by('name')
        extra_workers = Worker.objects.filter(cost_center = '74101')
        extra_extra_workers = Worker.objects.filter(personal_number=20000441)
        worker_objects = worker_objects.union(extra_workers, extra_extra_workers).order_by('name')

    if userObject.name in ['Pauchová Monika', 'Rejn Aleš', 'Pakan Michal']:
        extra_workers = Worker.objects.filter(cost_center = '64290').order_by('name')
        worker_objects = worker_objects.union(extra_workers).order_by('name')

    workers = {}

    for worker in worker_objects:
        workers[worker.personal_number] = []
        workers[worker.personal_number].append(worker.name)
        workers[worker.personal_number].append(worker.manager)
        workers[worker.personal_number].append(worker.costcenter.name)
        

    if request.method == 'POST':
        personal_number = request.POST.get('worker-number')
        worker = Worker.objects.get(personal_number=personal_number)
        shift = request.POST.get('shift-select')
        start_day = datetime.datetime.strptime(request.POST.get('date-of-availability'), '%Y-%m-%d')
        end_day = datetime.datetime.strptime(request.POST.get('date-of-availability-to'), '%Y-%m-%d')
        start_time = time.strptime(request.POST.get('start-time'), '%H:%M')
        start = datetime.datetime.combine(start_day, datetime.time(hour=start_time.tm_hour, minute=start_time.tm_min))
        end_time = time.strptime(request.POST.get('end-time'), '%H:%M')
        end = datetime.datetime.combine(end_day, datetime.time(hour=end_time.tm_hour, minute=end_time.tm_min))
        if not str(end - start)[-7:] == "8:00:00":
            standard = False
        else:
            standard = True

        while start_day <= end_day:
            start = datetime.datetime.combine(start_day, datetime.time(hour=start_time.tm_hour, minute=start_time.tm_min))
            end = datetime.datetime.combine(start_day, datetime.time(hour=end_time.tm_hour, minute=end_time.tm_min)) if end_time > start_time else datetime.datetime.combine(start_day + datetime.timedelta(days=1), datetime.time(hour=end_time.tm_hour, minute=end_time.tm_min))
            Offer.objects.create(worker=worker, day=start_day, shift=shift, created_by=request.user, start=start, end=end, standard=standard)
            start_day += datetime.timedelta(days=1)

        text = "Hotovo, pracovník byl přidán na burzu"
        messages.add_message(request, messages.INFO, text)


    if request.method == 'GET':
        pass

    context.update({
        'workers': workers,
        'selected_year': selected_year,
        'selected_month': f'{selected_month:02}',
        'selected_day': f'{selected_day:02}',
    })


    return render(request, 'add_offer.html', context)


@login_required
def make_request(request, selected_day=datetime.date.today().day, selected_month=datetime.date.today().month, selected_year=datetime.date.today().year):

    context = {}

    if request.user.is_authenticated:
        context['user_name'] = str(request.user.last_name) + ' ' + (request.user.first_name)

    userObject = Worker.objects.get(name = context['user_name'])
    
    if not userObject.is_manager:
        users_manager = Worker.objects.get(name=userObject.manager)
        production_units = Worker.objects.filter(manager__in=[users_manager, userObject.name]).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name')
    
    if userObject.is_manager:
        production_units = Worker.objects.filter(manager = userObject.name).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name')

    if request.method == 'POST':

        shift = request.POST.get('shift-select')
        start_day = datetime.datetime.strptime(request.POST.get('requested-date'), '%Y-%m-%d')
        end_day = datetime.datetime.strptime(request.POST.get('requested-date-to'), '%Y-%m-%d')
        target_cost_center = request.POST.get('production-unit-select')
        target_cost_center_object = CostCenter.objects.filter(name=target_cost_center).first()
        start_time = time.strptime(request.POST.get('start-time'), '%H:%M')
        start = datetime.datetime.combine(start_day, datetime.time(hour=start_time.tm_hour, minute=start_time.tm_min))
        end_time = time.strptime(request.POST.get('end-time'), '%H:%M')
        end = datetime.datetime.combine(end_day, datetime.time(hour=end_time.tm_hour, minute=end_time.tm_min))
        if (end - start) >= datetime.timedelta(hours=10):
            try:
                dtime = datetime.datetime.strptime(str(end - start)[-8:], "%H:%M:%S")
            except:
                dtime = datetime.datetime.strptime(str(end - start)[-7:], "%H:%M:%S")
        else:
            try:
                dtime = datetime.datetime.strptime(str(end - start)[-7:], "%H:%M:%S")
            except:
                dtime = datetime.datetime.strptime(str(end - start)[-8:], "%H:%M:%S")

        count_times = int(request.POST.get('count-of-workers'))

        if not str(end - start)[-7:] == "8:00:00":
            standard = False
        else:
            standard = True

        while start_day <= end_day:
            for _ in range(count_times):
                Request.objects.create(day=start_day, shift=shift, target_costcenter=target_cost_center_object, created_by=request.user, start=start, end=start + datetime.timedelta(hours=dtime.hour, minutes=dtime.minute), standard=standard)
            start_day += datetime.timedelta(days=1)
            if start_day == end_day and shift == "Noční":
                break
            start += datetime.timedelta(days=1)

        text = "Hotovo, poptávka pracovníka byla umístěna na burzu"
        messages.add_message(request, messages.INFO, text)

    if request.method == 'GET':
        pass

    context = {
        'selected_year': selected_year,
        'selected_month': f'{selected_month:02}',
        'selected_day': f'{selected_day:02}',
        'production_units': production_units,
    }

    return render(request, 'make_request.html', context)


@login_required
def transfer_from_offer(request, offer_id):

    context = {}

    if request.user.is_authenticated:

        userObject = Worker.objects.get(name = request.user.last_name + ' ' + request.user.first_name)
        
        context['user_name'] = str(request.user.last_name) + ' ' + (request.user.first_name)

        if userObject.is_manager:
            cost_centers = Worker.objects.filter(manager=context['user_name']).values_list('costcenter__name', 'costcenter__number').distinct('costcenter__name', 'costcenter__number')
        if not userObject.is_manager:
            cost_centers = Worker.objects.filter(manager=userObject.manager).values_list('costcenter__name', 'costcenter__number').distinct('costcenter__name', 'costcenter__number')

    offerObject = Offer.objects.filter().get(id=offer_id)
    workerObject = offerObject.worker

    if request.method == 'GET':
        if (str(workerObject.manager) == context['user_name'] or offerObject.created_by == request.user) and not offerObject.created_by_id == 5:
            context['is_own'] = True
        else:
            context['is_own'] = False

    if request.method == 'POST':
        if request.POST['btnSubmit'] == 'Smazat':
            offerObject.delete()
            text = 'Nabídka byla smazána'
            messages.add_message(request, messages.INFO, text)

        else:
            if offerObject.fullfiled == False:
                target_cost_center = request.POST.get('target_cost_center')
                target_cost_center_object = CostCenter.objects.get(number=target_cost_center)
                Assignment.objects.update_or_create(worker=workerObject, offer=offerObject, target_costcenter=target_cost_center_object, created_by=request.user)
                offerObject.fullfiled = True
                offerObject.save()

                assignmentObject = Assignment.objects.get(offer=offerObject)

                text = 'Výborně!  ' + str(workerObject.name) + ' se bude dne ' + str(offerObject.day) + ' hlásit na PU ' + str(assignmentObject.target_costcenter.name) + ' .'
                messages.add_message(request, messages.INFO, text)

                burza_user = BurzaUser.objects.get(burza_user=offerObject.created_by)

                if burza_user.notifications:
                    sendmail(text, "Nabídka byla využita", assignmentObject.offer.created_by.email, cc='peter.vajda@knorr-bremse.com')

    context.update({
        "worker": workerObject,
        "offer": offerObject,
        "cost_centers": cost_centers,
    })

    return render(request, 'transfer_from_offer.html', context)

@login_required
def transfer_from_request(request, request_id):

    context = {}

    if request.user.is_authenticated:
        context['user_name'] = str(request.user.last_name) + ' ' + (request.user.first_name)
        context['user_cost_center'] = list(Worker.objects.filter(manager=context['user_name']).values_list('costcenter__name', flat=True))

    userObject = Worker.objects.get(name=context['user_name'])

    if not userObject.is_manager:
        userObject_manager = Worker.objects.get(name=userObject.manager)
    else:
        userObject_manager = Worker.objects.get(name=context['user_name'])

    if userObject.is_manager:
        worker_objects = Worker.objects.filter(manager=context['user_name'], left=False).order_by('name')
    else:
        worker_objects = Worker.objects.filter(manager__in=[userObject_manager.name, userObject.name], left=False).order_by('name')

    if userObject.name == 'Szentkirályi Martin':
        worker_objects = Worker.objects.filter(manager = context['user_name'], not_available = False, left = False).order_by('name')
        extra_workers = Worker.objects.filter(cost_center = '74101')
        extra_extra_workers = Worker.objects.filter(personal_number=20000441)
        worker_objects = worker_objects.union(extra_workers, extra_extra_workers).order_by('name')

    if userObject.name in ['Pauchová Monika', 'Rejn Aleš', 'Pakan Michal']:
        extra_workers = Worker.objects.filter(cost_center = '64290').order_by('name')
        worker_objects = worker_objects.union(extra_workers).order_by('name')

    workers = {}

    for worker in worker_objects:
        workers[worker.personal_number] = []
        workers[worker.personal_number].append(worker.name)
        workers[worker.personal_number].append(worker.manager)
        workers[worker.personal_number].append(worker.costcenter.name)

    requestObject = Request.objects.get(id=request_id)

    time_from = requestObject.start.strftime("%H:%M")
    time_to = requestObject.end.strftime("%H:%M")

    context.update({
        'time_from': time_from,
        'time_to': time_to,
    })

    if request.method == 'GET':
        if str(requestObject.target_costcenter.name) in context['user_cost_center'] or requestObject.created_by == request.user:
            context['is_own'] = True
        else:
            context['is_own'] = False

    if request.method == 'POST':

        if request.POST['btnSubmit'] == 'Smazat poptávku':
            requestObject.delete()
            text = 'Poptávka byla smazána'
            messages.add_message(request, messages.INFO, text)

        else:
            worker_number = request.POST.get('worker-number')
            workerObject = Worker.objects.get(personal_number=worker_number)

            if requestObject.fullfiled is False:
                Assignment.objects.create(worker=workerObject, request=requestObject, target_costcenter=requestObject.target_costcenter, created_by=request.user)
                requestObject.fullfiled = True
                requestObject.save()

                assignmentObject = Assignment.objects.get(request=requestObject)

                
                if time.strptime(request.POST.get('time_to'), '%H:%M') < time.strptime(request.POST.get('time_from'), '%H:%M'):
                    duration = time.mktime(time.strptime(request.POST.get('time_to'), '%H:%M')) - time.mktime(time.strptime(request.POST.get('time_from'), '%H:%M'))
                    duration = (24*3600) + duration
                else:
                    duration = time.mktime(time.strptime(request.POST.get('time_to'), '%H:%M')) - time.mktime(time.strptime(request.POST.get('time_from'), '%H:%M'))
                
                assignmentObject.hours = duration / 3600
                assignmentObject.save()

                if not requestObject.shift == "Noční":
                    assignmentObject.assignment_start = datetime.datetime.strptime(str(requestObject.day) + ' ' + str(request.POST.get('time_from')), "%Y-%m-%d %H:%M")
                    assignmentObject.assignment_end = datetime.datetime.strptime(str(requestObject.day) + ' ' + str(request.POST.get('time_to')), "%Y-%m-%d %H:%M")
                    assignmentObject.save()
                else:
                    if time.mktime(time.strptime(request.POST.get('time_to'), '%H:%M')) < time.mktime(time.strptime(request.POST.get('time_from'), '%H:%M')):
                        assignmentObject.assignment_start = datetime.datetime.strptime(str(requestObject.day) + ' ' + str(request.POST.get('time_from')), "%Y-%m-%d %H:%M")
                        assignmentObject.assignment_end = datetime.datetime.strptime(str(requestObject.day) + ' ' + str(request.POST.get('time_to')), "%Y-%m-%d %H:%M") + datetime.timedelta(days=1)
                        assignmentObject.save()
                    else:
                        assignmentObject.assignment_start = datetime.datetime.strptime(str(requestObject.day) + ' ' + str(request.POST.get('time_from')), "%Y-%m-%d %H:%M") + datetime.timedelta(days=1)
                        assignmentObject.assignment_end = datetime.datetime.strptime(str(requestObject.day) + ' ' + str(request.POST.get('time_to')), "%Y-%m-%d %H:%M")  + datetime.timedelta(days=1)
                        assignmentObject.save()

                text = 'Výborně!  ' + str(workerObject.name) + ' se bude dne ' + str(requestObject.day) + ' hlásit na PU ' + str(assignmentObject.target_costcenter.name) + '.'
                messages.add_message(request, messages.INFO, text)

                if Offer.objects.filter(worker=workerObject, shift=requestObject.shift, day=requestObject.day).exists():

                    existingOffer = Offer.objects.get(worker=workerObject, shift=requestObject.shift, day=requestObject.day)
                    existingOffer.delete()

                burzaUser = BurzaUser.objects.get(burza_user=requestObject.created_by)

                if burzaUser.notifications:
                    sendmail(text, "Poptávka byla využita", requestObject.created_by.email, cc='peter.vajda@knorr-bremse.com')


    context.update({
        "request": requestObject,
        "workers": workers,
    })

    return render(request, 'transfer_from_request.html', context)


@login_required
def show_assignments(request):

    context = {}

    if request.user.is_authenticated:
        context['user_name'] = str(request.user.last_name) + ' ' + (request.user.first_name)
    
    userObject = Worker.objects.get(name = context['user_name'])

    if not userObject.is_manager:
        production_units = Worker.objects.filter(manager__in=[userObject.name, userObject.manager]).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name')

    if userObject.is_manager:
        production_units = Worker.objects.filter(manager = userObject.name).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name')
        
    Assignments1 = Assignment.objects.filter(target_costcenter__name__in=production_units).order_by('offer__day')
    Assignments2 = Assignment.objects.filter(offer__worker__costcenter__name__in=production_units).order_by('offer__day')
    Assignments3 = Assignment.objects.filter(request__target_costcenter__name__in=production_units).order_by('offer__day')
    Assignments = Assignments1.union(Assignments2, Assignments3)

    context.update({
        "assignments": Assignments,
    })

    return render(request, 'show_assignments.html', context)


@login_required
def show_offers(request):

    context = {}

    if request.user.is_authenticated:
        context['user_name'] = str(request.user.last_name) + ' ' + (request.user.first_name)
    
    userObject = Worker.objects.get(name = context['user_name'])

    if not userObject.is_manager:
        production_units = Worker.objects.filter(manager = userObject.manager).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name')

    if userObject.is_manager:
        production_units = Worker.objects.filter(manager = userObject.name).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name')

    offers = Offer.objects.filter(worker__costcenter__name__in=production_units).order_by('day')
    assignments = Assignment.objects.all()

    for offer in offers:
        offer.used = False
        if assignments.filter(offer=offer).count() > 0:
            offer.used = True
            if assignments.filter(offer=offer, realised=True).count() > 0:
                offer.realised = True
            else:
                offer.realised = False

    context = {
        "offers": offers,
    }

    return render(request, 'show_offers.html', context)


@login_required
def show_requests(request):

    context = {}

    if request.user.is_authenticated:
        context['user_name'] = str(request.user.last_name) + ' ' + (request.user.first_name)
    
    userObject = Worker.objects.get(name = context['user_name'])

    if not userObject.is_manager:
        production_units = Worker.objects.filter(manager = userObject.manager).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name')

    if userObject.is_manager:
        production_units = Worker.objects.filter(manager = userObject.name).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name')

    requests = Request.objects.filter(target_costcenter__name__in=production_units).order_by('day')
    assignments = Assignment.objects.all()

    for req in requests:
        req.used = False
        if assignments.filter(request=req).count() > 0:
            req.used = True
            if assignments.filter(request=req, realised=True).count() > 0:
                req.realised = True
            else:
                req.realised = False

    context = {
        "requests": requests,
    }

    return render(request, 'show_requests.html', context)

@login_required
def edit_assignments(request):

    context = {}

    if request.user.is_authenticated:
        context['user_name'] = str(request.user.last_name) + ' ' + (request.user.first_name)
    
    userObject = Worker.objects.get(name = context['user_name'])

    if not userObject.is_manager:
        production_units = list(Worker.objects.filter(manager = userObject.manager).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name'))

    if userObject.is_manager:
        production_units = list(Worker.objects.filter(manager = userObject.name).values_list('costcenter__name', flat=True).distinct().order_by('costcenter__name'))

    if request.method == 'POST':

        filters = {}

        if request.POST.get('date_from'):
            date_from = request.POST.get('date_from')
            context.update({'date_from': date_from})

        if request.POST.get('date_to'):
            date_to = request.POST.get('date_to')
            context.update({'date_to': date_to})

        if not request.POST.get('date_to') and not request.POST.get('date_from'):
            date_to = datetime.date.today()
            date_from = datetime.date.today()

        if request.POST.get('edit'):
            for ident in request.POST.getlist('identifier'):
                Assignment.objects.filter(id=ident).update(hours=request.POST.getlist('duration')[request.POST.getlist('identifier').index(ident)])
                if str('realised-' + ident) in request.POST.dict():
                    Assignment.objects.filter(id=ident).update(realised=True)
                else:
                    Assignment.objects.filter(id=ident).update(realised=False)
            if request.POST.get('date_to'):
                filters['offer__day__lte'] = request.POST.get('date_to')
            if request.POST.get('date_from'):
                filters['offer__day__gte'] = request.POST.get('date_from')

            if request.POST.get('duration'):
                Assignment.objects.filter(id=ident).update(hours=request.POST.get('duration'))

            text = 'Změny uloženy.'
            messages.add_message(request, messages.INFO, text)

        Assignments = Assignment.objects.filter(Q(offer__day__lte=date_to) | Q(request__day__lte=date_to)).filter(Q(offer__day__gte=date_from) | Q(request__day__gte=date_from)).filter(target_costcenter__name__in=production_units)

        context.update({'assignments': Assignments})

    if request.method == 'GET':
        pass

    return render(request, 'edit_assignments.html', context)


def show_charts(request):

    context = {}
    calendar_to_web = {}
    week_num = 0

    cal = calendar.Calendar().yeardatescalendar(2023 , 1)
    
    for month in cal:
        calendar_to_web[cal.index(month)] = {}
        for useless_list in month:
            for week in useless_list:
                calendar_to_web[cal.index(month)][week_num] = week
                week_num += 1

    context['offersDataToChart'] = str()

    context['test'] = 'Show charts!'
    context['calendar'] = calendar_to_web

    return render(request, 'charts.html', context)


def sendmail(message_content, subject, recipient, cc='peter.vajda@knorr-bremse.com'):

    from email.message import EmailMessage

    msg = EmailMessage()

    msg['Subject'] = subject
    msg['From'] = "automat@knorr-bremse.com"
    msg['To'] = recipient
    msg['Cc'] = cc
    msg.set_content(message_content, subtype='html')

    s = smtplib.SMTP('smtp-relay.corp.knorr-bremse.com')
    s.send_message(msg)
    s.quit()


@login_required
def employees(request):

    manager = request.user.last_name + ' ' + request.user.first_name

    employees = Worker.objects.filter(manager = manager, left=False).order_by('name')

    keys = []
    values = []

    if request.method == 'POST':
        for k, v in dict(request.POST).items():
            keys.append(k)

        keys.pop(0)
        keys = keys[:-1]

        for personal_number in keys:
            if not personal_number[0] == 'r':
                Worker.objects.filter(personal_number=personal_number).update(not_available=True)
            else:
                Worker.objects.filter(personal_number=personal_number[9:]).update(not_available=False)

    context = {
        "employees": employees,
    }
    if request.method == "GET":
        return render(request, 'employees.html', context)

    if request.method == "POST":
        return HttpResponseRedirect('/burza')
        

def employee_details(request, personal_number):
    
    workerobject = Worker.objects.get(personal_number=personal_number)
    offers = Offer.objects.filter(worker=workerobject)
    assignments = Assignment.objects.filter(worker=workerobject)

    context = {
        "offers": offers,
        "assignments": assignments,
        "today": datetime.date.today(),
        "personal_number": personal_number,
    }

    return render(request, 'employee_detail.html', context)


@login_required
def cancel_offer(request, offer_id):

    offerObject = Offer.objects.filter(id=offer_id)
    offerObject.delete()

    return HttpResponseRedirect(request.GET.get('next'))
    

@login_required
def cancel_request(request, request_id):

    requestObject = Request.objects.filter(id=request_id)
    requestObject.delete()

    return HttpResponseRedirect(request.GET.get('next'))


@login_required
def agree_to_cancel(request, pending_request_id):

    pending_requestObject = PendingRequest.objects.get(id=pending_request_id)
    assignmentObject = Assignment.objects.get(pendingrequest=pending_requestObject)
    message_content = ""

    try:
        offerObject = Offer.objects.get(assignment__pendingrequest=pending_requestObject)
        worker_that_created_offer = Worker.objects.get(name=offerObject.created_by.last_name + ' ' + offerObject.created_by.first_name)
        worker_that_created_pending_request = Worker.objects.get(name=pending_requestObject.created_by.last_name + ' ' + pending_requestObject.created_by.first_name)
        message_content += offerObject.worker.name
        message_content += " - "
        message_content += str(offerObject.day)
        message_content += " - zrušeno."

        if worker_that_created_offer.is_manager:
            worker_that_created_offer_manager = worker_that_created_offer
        else:
            worker_that_created_offer_manager = worker_that_created_offer.manager

        if worker_that_created_pending_request.is_manager:
            worker_that_created_pending_request_manager = worker_that_created_pending_request
        else:
            worker_that_created_pending_request_manager = worker_that_created_pending_request.manager

        
        if worker_that_created_offer_manager == worker_that_created_pending_request_manager:
            offerObject.delete()
        else:
            offerObject.fullfiled = False
            offerObject.save()
    except:
        pass

    try:
        requestObject = Request.objects.get(assignment__pendingrequest=pending_requestObject)
        message_content += str(requestObject.day)
        message_content += " - zrušeno."
        worker_that_created_request = Worker.objects.get(name=requestObject.created_by.last_name + ' ' + requestObject.created_by.first_name)
        worker_that_created_pending_request = Worker.objects.get(name=pending_requestObject.created_by.last_name + ' ' + pending_requestObject.created_by.first_name)

        if worker_that_created_request.is_manager:
            worker_that_created_request_manager = worker_that_created_request
        else:
            worker_that_created_request_manager = worker_that_created_request.manager

        if worker_that_created_pending_request.is_manager:
            worker_that_created_pending_request_manager = worker_that_created_pending_request
        else:
            worker_that_created_pending_request_manager = worker_that_created_pending_request.manager

        
        if worker_that_created_request_manager == worker_that_created_pending_request_manager:
            requestObject.delete()
        else:
            requestObject.fullfiled = False
            requestObject.save()
    except:
        pass

    pending_requestObject.delete()

    assignmentObject.delete()

    sendmail(message_content, "Souhlasím se zrušením zápujčky", pending_requestObject.created_by.email)

    return HttpResponseRedirect(request.GET.get('next'))


@login_required
def decline_to_cancel(request, pending_request_id):

    pending_requestObject = PendingRequest.objects.get(id=pending_request_id)

    pending_requestObject.declined = True
    pending_requestObject.save()

    message_content = 'Nesouhlasím se zrušením zápujčky ' + str(pending_requestObject.assignment.offer.day) + ' - ' + pending_requestObject.assignment.offer.shift + ' směna - ' + pending_requestObject.assignment.offer.worker.name
    recipient = str(pending_requestObject.assignment.offer.worker.manager)
    recipient_first_name = recipient.split(' ')[1]
    recipient_last_name = recipient.split(' ')[0]
    recipientObject = User.objects.get(first_name=recipient_first_name, last_name=recipient_last_name)

    sendmail(message_content, "Nesouhlasím se zrušením zápujčky", recipientObject.email)

    return HttpResponseRedirect(request.GET.get('next'))


@login_required
def dashboard(request):

    context = {}

    user = User.objects.get(last_name=request.user.last_name, first_name=request.user.first_name)
    user_worker = Worker.objects.get(name=request.user.last_name + ' ' + request.user.first_name)

    if not user_worker.is_manager:
        my_cost_centers = list(Worker.objects.filter(manager=user_worker.manager).values_list('costcenter__name', flat=True).distinct('costcenter__name'))
        workers = list(Worker.objects.filter(manager=user_worker.manager))
    else:
        my_cost_centers = list(Worker.objects.filter(manager=user_worker.name).values_list('costcenter__name', flat=True).distinct('costcenter__name'))
        workers = list(Worker.objects.filter(manager=user_worker.name))
    
    manager = user_worker.name

    allPendingRequests = PendingRequest.objects.all()

    pendingRequestsMyWorkers = []
    pendingRequestsMyApproval = []

    for pendingRequest in allPendingRequests:

        worker_that_created_pending_request = Worker.objects.get(name=pendingRequest.created_by.last_name + ' ' + pendingRequest.created_by.first_name)
        worker_that_created_assignment = Worker.objects.get(name=pendingRequest.assignment.created_by.last_name + ' ' + pendingRequest.assignment.created_by.first_name)
        try:
            worker_that_created_offer = Worker.objects.get(name=pendingRequest.assignment.offer.created_by.last_name + ' ' + pendingRequest.assignment.offer.created_by.first_name)
        except:
            worker_that_created_request = Worker.objects.get(name=pendingRequest.assignment.request.created_by.last_name + ' ' + pendingRequest.assignment.request.created_by.first_name)

        try:
            if worker_that_created_request.costcenter.name in my_cost_centers and not worker_that_created_pending_request.costcenter.name in my_cost_centers:
                pendingRequestsMyApproval.append(pendingRequest)
            elif not worker_that_created_request.costcenter.name in my_cost_centers and worker_that_created_pending_request.costcenter.name in my_cost_centers:
                pendingRequestsMyWorkers.append(pendingRequest)
            elif not worker_that_created_request.costcenter.name in my_cost_centers and worker_that_created_assignment.costcenter.name in my_cost_centers:
                pendingRequestsMyApproval.append(pendingRequest)
    
        except:
            if worker_that_created_offer.costcenter.name in my_cost_centers and not worker_that_created_pending_request.costcenter.name in my_cost_centers:
                pendingRequestsMyApproval.append(pendingRequest)
            elif not worker_that_created_offer.costcenter.name in my_cost_centers and worker_that_created_pending_request.costcenter.name in my_cost_centers:
                pendingRequestsMyWorkers.append(pendingRequest)
            elif not worker_that_created_offer.costcenter.name in my_cost_centers and worker_that_created_assignment in my_cost_centers:
                pendingRequestsMyApproval.append(pendingRequest)

    activeTakenOffers = Assignment.objects.filter(target_costcenter__name__in=my_cost_centers, offer__day__gte=datetime.date.today() - datetime.timedelta(days=2)).union(Assignment.objects.filter(target_costcenter__name__in=my_cost_centers, request__day__gte=datetime.date.today())).order_by('assignment_start')
    activeUsedOffersFromMyCenters = Assignment.objects.filter(worker__in=workers, offer__day__gte=datetime.date.today() - datetime.timedelta(days=2)).union(Assignment.objects.filter(worker__in=workers, request__day__gte=datetime.date.today()))
    activeUsedOffersFromMyCenters = activeUsedOffersFromMyCenters.order_by('assignment_start', 'worker')
    myUntakenOffers = Offer.objects.filter(fullfiled=False, worker__in=workers, day__gte=datetime.date.today() - datetime.timedelta(days=2)).order_by('start')
    myUnusedRequests = Request.objects.filter(fullfiled=False, target_costcenter__name__in=my_cost_centers, day__gte=datetime.date.today() - datetime.timedelta(days=2)).order_by('start')

    context['pending_requests_my_workers'] = pendingRequestsMyWorkers
    context['pending_requests_my_approval'] = pendingRequestsMyApproval
    context['active_taken_offers'] = activeTakenOffers
    context['active_used_offers_from_my_centers'] = activeUsedOffersFromMyCenters
    context['my_untaken_offers'] = myUntakenOffers
    context['my_unused_requests'] = myUnusedRequests
    context['manager'] = manager
    context['cost_centers'] = my_cost_centers

    return render(request, 'dashboard.html', context)

@login_required
def make_request_to_cancel_assignment(request, assignment_id):

    user_worker = Worker.objects.get(name=request.user.last_name + ' ' + request.user.first_name)
    user = User.objects.get(last_name=request.user.last_name, first_name=request.user.first_name)
    
    if not user_worker.is_manager:
        users_manager = Worker.objects.get(name=user_worker.manager)
    else:
        users_manager = user_worker
    
    users_manager_workers = Worker.objects.filter(manager=users_manager)

    message_content = ''

    assignmentObject = Assignment.objects.get(id=assignment_id)

    pending_requestObject = PendingRequest.objects.create(assignment=assignmentObject, created_by=user)

    assignmentObject.pending_cancellation = True
    assignmentObject.save()

    worker_that_created_assignment = Worker.objects.get(name=assignmentObject.created_by.last_name + ' ' + assignmentObject.created_by.first_name)
    
    try:
        worker_that_created_request = Worker.objects.get(name=assignmentObject.request.created_by.last_name + ' ' + assignmentObject.request.created_by.first_name)
    except:
        worker_that_created_offer = Worker.objects.get(name=assignmentObject.offer.created_by.last_name + ' ' + assignmentObject.offer.created_by.first_name)

    
    if not pending_requestObject.created_by == assignmentObject.created_by:
        recipientObject = assignmentObject.created_by
    else:
        try:
            recipientObject = assignmentObject.offer.created_by
        except:
            recipientObject = assignmentObject.request.created_by

    message_content += '<a href="10.49.34.115/burza/pending_request/' + str(pending_requestObject.id) + '/agree//?next=/burza">Souhlasím</a>'
    message_content += '<br /><br />'
    message_content += '<a href="10.49.34.115/burza/pending_request/' + str(pending_requestObject.id) + '/decline//?next=/burza">Nesouhlasím</a>'

    try:
        sendmail(message_content, "Žádost o zrušení zápujčky " + str(assignmentObject.request.day) + ' - ' + assignmentObject.request.shift + ' - ' + assignmentObject.worker.name, recipientObject.email)
    except:
        sendmail(message_content, "Žádost o zrušení zápujčky " + str(assignmentObject.offer.day) + ' - ' + assignmentObject.offer.shift + ' - ' + assignmentObject.worker.name, recipientObject.email)


    return HttpResponseRedirect(request.GET.get('next'))

    
def visuals(request):

    if request.method == "GET":

        offers = Offer.objects.values('day').annotate(countie=Count('day')).order_by('day')
        assignments = Assignment.objects.all()
        requests = Request.objects.all().order_by('day')
        unused_offers = Offer.objects.filter(fullfiled=False)
        unused_requests = Request.objects.filter(fullfiled=False)

        pie_chart_offers = []
        pie_chart_requests = []
        data_dict_offers = {}
        data_dict_requests = {}
        data_dict_offers['values'] = []
        data_dict_offers['values'].extend([unused_offers.count(), (offers.count() - unused_offers.count())])
        data_dict_offers['labels'] = []
        data_dict_offers['labels'].extend(['Nevyužité nabídky', 'Využité nabídky'])

        data_dict_requests['values'] = []
        data_dict_requests['values'].extend([unused_requests.count(), (requests.count() - unused_requests.count())])
        data_dict_requests['labels'] = []
        data_dict_requests['labels'].extend(['Nevyužité poptávky', 'Využité poptávky'])

        pie_chart_offers.append(data_dict_offers)
        pie_chart_requests.append(data_dict_requests)

        plotly_offers_x = []
        plotly_offers_y = []
        plotly_requests_y = []
        plotly_requests_x = []
        plotly_assignments_y = []

        for offer in offers:
            plotly_offers_x.append(str(offer['day']))
            plotly_offers_y.append(offer['countie'])

        for r in requests:
            plotly_requests_x.append(str(r.day))

        plotly_requests_x = list(set(plotly_requests_x))
        plotly_requests_x.sort()

        range_of_x = max((datetime.datetime.strptime(plotly_offers_x[-1], '%Y-%m-%d') - datetime.datetime.strptime(plotly_offers_x[0], '%Y-%m-%d')).days, (datetime.datetime.strptime(plotly_requests_x[-1], '%Y-%m-%d') - datetime.datetime.strptime(plotly_requests_x[0], '%Y-%m-%d')).days) 
        base_date = min(datetime.datetime.strptime(plotly_offers_x[0], '%Y-%m-%d').date(), datetime.datetime.strptime(plotly_requests_x[0], '%Y-%m-%d').date())
        list_of_dates = [base_date + datetime.timedelta(days=x) for x in range(range_of_x)]

        for date in list_of_dates:
            if not str(date) in plotly_offers_x:
                plotly_offers_x.append(str(date))
                plotly_offers_y.append(0)
            plotly_requests_y.append(requests.filter(day=date).count())
        
        new_plotly_offers_x = []
        new_plotly_offers_y = []
        
        for item in list(sorted(zip(plotly_offers_x, plotly_offers_y))):
            new_plotly_offers_x.append(item[0])
            new_plotly_offers_y.append(item[1])

        plotly_x_beginning = plotly_offers_x[0]
        plotly_x_end = plotly_offers_x[-1]

    if request.method == "POST":

        filters = {}

        if request.POST.get('date_from'):
            filters['day__gte'] = request.POST.get('date_from')
        else:
            filters['day__gte'] = datetime.date.today() - datetime.timedelta(days=30)

        if request.POST.get('date_to'):
            filters['day__lte'] = request.POST.get('date_to')
        else:
            filters['day__lte'] = datetime.date.today()

        offers = Offer.objects.filter(**filters).values('day').annotate(countie=Count('day')).order_by('day')
        requests = Request.objects.filter(**filters).order_by('day')
        unused_offers = Offer.objects.filter(fullfiled=False).filter(**filters)
        unused_requests = Request.objects.filter(fullfiled=False).filter(**filters)

        pie_chart_offers = []
        pie_chart_requests = []
        data_dict_offers = {}
        data_dict_requests = {}
        data_dict_offers['values'] = []
        data_dict_offers['values'].extend([unused_offers.count(), (offers.count() - unused_offers.count())])
        data_dict_offers['labels'] = []
        data_dict_offers['labels'].extend(['Nevyužité nabídky', 'Využité nabídky'])

        data_dict_requests['values'] = []
        data_dict_requests['values'].extend([unused_requests.count(), (requests.count() - unused_requests.count())])
        data_dict_requests['labels'] = []
        data_dict_requests['labels'].extend(['Nevyužité poptávky', 'Využité poptávky'])

        pie_chart_offers.append(data_dict_offers)
        pie_chart_requests.append(data_dict_requests)

        plotly_offers_x = []
        plotly_offers_y = []
        plotly_requests_y = []
        plotly_requests_x = []
        plotly_assignments_y = []

        for offer in offers:
            plotly_offers_x.append(str(offer['day']))
            plotly_offers_y.append(offer['countie'])

        for r in requests:
            plotly_requests_x.append(str(r.day))

        plotly_requests_x = list(set(plotly_requests_x))
        plotly_requests_x.sort()

        range_of_x = max((datetime.datetime.strptime(plotly_offers_x[-1], '%Y-%m-%d') - datetime.datetime.strptime(plotly_offers_x[0], '%Y-%m-%d')).days, (datetime.datetime.strptime(plotly_requests_x[-1], '%Y-%m-%d') - datetime.datetime.strptime(plotly_requests_x[0], '%Y-%m-%d')).days) 
        base_date = min(datetime.datetime.strptime(plotly_offers_x[0], '%Y-%m-%d').date(), datetime.datetime.strptime(plotly_requests_x[0], '%Y-%m-%d').date())
        list_of_dates = [base_date + datetime.timedelta(days=x) for x in range(range_of_x)]

        for date in list_of_dates:
            if not str(date) in plotly_offers_x:
                plotly_offers_x.append(str(date))
                plotly_offers_y.append(0)
            plotly_requests_y.append(requests.filter(day=date).count())
            plotly_assignments_y.append(Assignment.objects.filter(Q(request__day=date) | Q(offer__day=date)).count())

        
        new_plotly_offers_x = []
        new_plotly_offers_y = []
        
        for item in list(sorted(zip(plotly_offers_x, plotly_offers_y))):
            new_plotly_offers_x.append(item[0])
            new_plotly_offers_y.append(item[1])

        plotly_x_beginning = plotly_offers_x[0]
        plotly_x_end = plotly_offers_x[-1]
        

    context = {
        'plotly_offers_x': new_plotly_offers_x,
        'plotly_offers_y': new_plotly_offers_y,
        'plotly_requests_y': plotly_requests_y,
        'plotly_x_beginning': plotly_x_beginning,
        'plotly_x_end': plotly_x_end,
        'pie_chart_offers': pie_chart_offers,
        'pie_chart_requests': pie_chart_requests,
    }

    if request.method == "POST":

        context.update({
            'date_from': filters['day__gte'],
            'date_to': filters['day__lte'],
        })

    
    plotly_assignments_x = {(datetime.datetime.strptime(xdate, '%Y-%m-%d').replace(day=1)).strftime("%Y-%m-%d") for xdate in plotly_offers_x}
    plotly_assignments_y = [Assignment.objects.filter(Q(request__day__gte=date, request__day__lte=datetime.datetime.strptime(date, '%Y-%m-%d') + datetime.timedelta(days=30)) | Q(offer__day__gte=date, offer__day__lte=datetime.datetime.strptime(date, '%Y-%m-%d') + datetime.timedelta(days=30))).count() for date in plotly_assignments_x]

    context.update({
        'plotly_assignments_x': list(plotly_assignments_x),
        'plotly_assignments_y': plotly_assignments_y,
    })

    return render(request, 'visuals.html', context)


def notifications_switch(request):

    burza_user = BurzaUser.objects.get(burza_user=request.user)

    if burza_user.notifications:
        burza_user.notifications = False
        burza_user.save()
    else:
        burza_user.notifications = True
        burza_user.save()

    return HttpResponseRedirect('/burza')


def help(request):

    context = {}

    return render(request, 'help.html', context)


def reports(request):

    filters = {}

    if request.method == "GET":
        filters['day__gte'] = datetime.date.today() - datetime.timedelta(days=90)
        filters['day__lte'] = datetime.date.today()

    if request.method == "POST":
        if request.POST.get('date_from'):
            filters['day__gte'] = request.POST.get('date_from')
        else:
            filters['day__gte'] = datetime.date.today() - datetime.timedelta(days=90)

        if request.POST.get('date_to'):
            filters['day__lte'] = request.POST.get('date_to')
        else:
            filters['day__lte'] = datetime.date.today()

    as1 = Assignment.objects.all().exclude(request=None).filter(request__day__gte=filters['day__gte'])
    as2 = Assignment.objects.all().exclude(offer=None).filter(offer__day__gte=filters['day__gte'])
    asA = as1.union(as2)
    as3 = Assignment.objects.all().exclude(request=None).filter(request__day__lte=filters['day__lte'])
    as4 = Assignment.objects.all().exclude(offer=None).filter(offer__day__lte=filters['day__lte'])
    asB = as3.union(as4)

    assignments = asA.intersection(asB)

    requests = Request.objects.all().filter(**filters)
    offers = Offer.objects.all().filter(**filters).filter(fullfiled=True)


    req_dict = {}

    for rq in requests:
        if not rq.target_costcenter_id in req_dict:
            req_dict[rq.target_costcenter_id] = {}
            req_dict[rq.target_costcenter_id]['wanted'] = (rq.end - rq.start).total_seconds() / 3600
            if (rq.end - rq.start).total_seconds() / 3600 > 7.5:
                req_dict[rq.target_costcenter_id]['wanted'] -= 0.5
            try:
                req_dict[rq.target_costcenter_id]['got'] = Assignment.objects.values_list('hours', flat=True).get(request_id=rq.id) if Assignment.objects.values_list('hours', flat=True).get(request_id=rq.id) < 8 else Assignment.objects.values_list('hours', flat=True).get(request_id=rq.id) - 0.5
            except:
                req_dict[rq.target_costcenter_id]['got'] = 0 
        else:
            if rq.fullfiled:
                try:
                    req_dict[rq.target_costcenter_id]['wanted'] += (rq.end - rq.start).total_seconds() / 3600 if (rq.end - rq.start).total_seconds() / 3600 < 8 else (rq.end - rq.start).total_seconds() / 3600 - 0.5
                    req_dict[rq.target_costcenter_id]['got'] += Assignment.objects.values_list('hours', flat=True).get(request_id=rq.id) if Assignment.objects.values_list('hours', flat=True).get(request_id=rq.id) < 8 else Assignment.objects.values_list('hours', flat=True).get(request_id=rq.id) - 0.5
                except:
                    pass
            else:
                req_dict[rq.target_costcenter_id]['wanted'] += (rq.end - rq.start).total_seconds() / 3600 if (rq.end - rq.start).total_seconds() / 3600 < 8 else (rq.end - rq.start).total_seconds() / 3600 - 0.5


    for of in offers:
        assig = Assignment.objects.get(offer_id=of.id)

        if of.fullfiled:
            if not assig.target_costcenter_id in req_dict:
                req_dict[assig.target_costcenter_id] = {}
                req_dict[assig.target_costcenter_id]['wanted'] = 0
                req_dict[assig.target_costcenter_id]['got'] = assig.hours if assig.hours < 8 else assig.hours - 0.5
            else:
                req_dict[assig.target_costcenter_id]['wanted'] += assig.hours if assig.hours < 8 else assig.hours - 0.5
                req_dict[assig.target_costcenter_id]['got'] += assig.hours if assig.hours < 8 else assig.hours - 0.5

    assignments_by_month = {}
    requests_by_month = {}

    assignments_months = []
    assignments_numbers = []
    assignments_months_named = []

    requests_numbers_wanted = []
    requests_numbers_got = []
    requests_months_named = []

    for assig in Assignment.objects.all():
        try:
            assig_date = assig.request.day
        except:
            try:
                assig_date = assig.offer.day
            except:
                pass
        

        if str(assig_date.month) + ' ' + str(assig_date.year) not in assignments_by_month:
            assignments_by_month[str(assig_date.month) + ' ' + str(assig_date.year)] = assig.hours if assig.hours < 8 else 7.5
        else:
            assignments_by_month[str(assig_date.month) + ' ' + str(assig_date.year)] += assig.hours if assig.hours < 8 else 7.5

    for month, number in assignments_by_month.items():
        assignments_months.append(month)
        assignments_numbers.append(number)
       
    for month in assignments_months:
        month_number = month.split()[0]
        month_year = month.split()[1]
        assignments_months_named.append(str(calendar.month_abbr[int(month_number)]) + ' ' + str(month_year))

    for req in Request.objects.all().order_by('day'):
        req_date = req.day

        if str(req_date.month) + ' ' + str(req_date.year) not in requests_by_month:
            requests_by_month[str(req_date.month) + ' ' + str(req_date.year)] = {}

    for string, value in requests_by_month.items():
        requests_by_month[string]['wanted'] = 0
        requests_by_month[string]['got'] = 0
        num_days = calendar.monthrange(int(string.split()[1]), int(string.split()[0]))[1]
        days_in_month = [datetime.date(int(string.split()[1]), int(string.split()[0]), day) for day in range(1, num_days+1)]

        all_reqs_in_month = Request.objects.filter(day__in=days_in_month)
        duration = 0
        for r in all_reqs_in_month:
            end = r.end
            start = r.start
            duration += (end - start).total_seconds() / 3600
        requests_by_month[string]['wanted'] = duration

        all_offs_in_month = Offer.objects.filter(day__in=days_in_month, fullfiled=True)
        duration = 0
        for o in all_offs_in_month:
            end = o.end
            start = o.start
            duration += (end - start).total_seconds() / 3600
        requests_by_month[string]['wanted'] += duration
            
        as1 = Assignment.objects.all().exclude(request=None).filter(request__day__in=days_in_month)
        as2 = Assignment.objects.all().exclude(offer=None).filter(offer__day__in=days_in_month)
        assignments2 = as1.union(as2)

        for assig in assignments2:
            requests_by_month[string]['got'] += ((assig.assignment_end - assig.assignment_start).total_seconds() / 3600) - 0.5 if assig.assignment_end else (assig.hours if assig.hours < 8 else 7.5)


    for key, value in requests_by_month.items():
        month_number = key.split()[0]
        month_year = key.split()[1]
        requests_numbers_wanted.append(value['wanted'])
        requests_numbers_got.append(value['got'])
        requests_months_named.append(str(calendar.month_abbr[int(month_number)]) + ' ' + str(month_year))

    if type(filters['day__lte']) == datetime.date:
        filters['day__lte'] = filters['day__lte'].strftime("%Y-%m-%d")

    if type(filters['day__gte']) == datetime.date:
        filters['day__gte'] = filters['day__gte'].strftime("%Y-%m-%d")

    context = {
        'assignments': assignments,
        'req_dict': req_dict,
        'date_to': filters['day__lte'],
        'date_from': filters['day__gte'],
        'assignments_months': assignments_months_named,
        'assignments_numbers': assignments_numbers,
        'requests_months': requests_months_named,
        'requests_wanted': requests_numbers_wanted,
        'requests_got': requests_numbers_got,
    }
    

    return render(request, 'reports.html', context)