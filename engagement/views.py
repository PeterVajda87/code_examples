from django.shortcuts import render, redirect
from .models import Questionnaire, QuestionnaireTemplate, Question, SurveyProfile, Section, RangedAnswer, TextAnswer, Employee
from burza.models import Worker
from django.http import JsonResponse
import datetime
from django.contrib.auth.decorators import login_required, permission_required
import json
from django.contrib.auth.models import User
from openpyxl import load_workbook
from django.core.exceptions import MultipleObjectsReturned



@login_required
def homepage_view(request):

    try:
        engagementUserObject = Employee.objects.get(name=str(request.user.last_name + ' ' + request.user.first_name))
    except:
        engagementUserObject = Employee.objects.filter(name=str(request.user.last_name + ' ' + request.user.first_name), is_supervisor=True).first()

    try:
        is_supervisor = engagementUserObject.is_supervisor
    except:
        is_supervisor = False

    if is_supervisor:
        return redirect('my_questionnaires_list')

    context = {
        'is_supervisor': is_supervisor,
    }

    return render(request, 'homepage.html', context) 


@login_required
def admin_view(request):

    if request.user.last_name in ['Kuštanová', 'Töröková', 'Vajda', 'Černá', 'Kubešová', 'Herdova']:

        context = {}
    
        return render(request, 'admin.html', context)

    else:

        return redirect('my_questionnaires_list')


def show_employees(request):

    employees = Employee.objects.all()
    supervisors = Employee.objects.filter(is_supervisor=True)

    context = {
        'employees': employees,
        'supervisors': supervisors,
    }

    return render(request, 'show_employees.html', context)


def add_employee(request):

    context = {}
    cost_centers = Employee.objects.all().values_list('cost_center_number', 'cost_center_name').distinct('cost_center_number')
    positions = Employee.objects.all().values_list('position', flat=True).distinct('position')

    if request.method == "GET":
        pass

    if request.method == "POST":
        Employee.objects.create(
            name=request.POST.get('name'),
            cost_center_number=request.POST.get('costcenter'),
            cost_center_name=Employee.objects.filter(cost_center_number=request.POST.get('costcenter')).values_list('cost_center_name', flat=True).first(),
            personal_number=request.POST.get('personal_number'),
            position=request.POST.get('position'),
            evaluator=request.POST.get('evaluator'),
            is_supervisor=bool(request.POST.get('supervisor', False)),
        )

        active_q_templates = QuestionnaireTemplate.objects.filter(locked=False)

        new_employee = Employee.objects.get(personal_number=request.POST.get('personal_number'))
        new_employee_evaluator = Employee.objects.get(is_supervisor=True, name=new_employee.evaluator)

        for q_template in active_q_templates:
            Questionnaire.objects.create(
                            interviewer = new_employee_evaluator,
                            questionnaire = q_template,
                            interviewee = new_employee,
                            )

    
    context.update({
        'cost_centers': cost_centers,
        'positions': positions,
        'evaluators': Employee.objects.filter(is_supervisor=True).values_list('name', flat=True).order_by('name').distinct().exclude(evaluator=None),
    })
    

    return render(request, 'add_employee.html', context)


def remove_employee(request, employee_id):

    employeeObject = Employee.objects.get(id=employee_id)
    employeeObject.delete()

    return JsonResponse({"answer": "ok"})


def edit_employee(request, employee_id):

    employeeObject = Employee.objects.get(id=employee_id)
    cost_centers = Employee.objects.all().values_list('cost_center_number', 'cost_center_name').distinct('cost_center_number')
    positions = Employee.objects.all().values_list('position', flat=True).distinct('position')

    if request.method == "POST":
        data = json.loads(request.POST['data'])
        employeeObject.name = data['name']
        employeeObject.personal_number = data['personal_number']
        employeeObject.cost_center_number = data['costcenter'].split(' - ')[0]
        employeeObject.cost_center_name = data['costcenter'].split(' - ')[1]
        employeeObject.position = data['position']
        employeeObject.evaluator = data['evaluator']
        employeeObject.is_supervisor = bool(data.get('supervisor', False))
        employeeObject.save()

        update_evaluators(None, employee_id)

        return JsonResponse({"answer": "ok"})

    context = {
        'employee': employeeObject,
        'cost_centers': cost_centers,
        'positions': positions,
        'evaluators': Employee.objects.filter(is_supervisor=True).values_list('name', flat=True).order_by('name').exclude(name=None),
    }

    return render(request, 'edit_employee.html', context)


@login_required
def create_questionnaire(request):

    question_order = 0

    surveyProfile_object = SurveyProfile.objects.get(user=request.user)
    ranged_answers = RangedAnswer.objects.all()

    if request.method == "POST":
        data = json.dumps(dict(request.POST))
        data = json.loads(data)

        for key, value in data.items():
            if key == "name":
                questionnaireTemplate_name = value
                new_questionnaire, created = QuestionnaireTemplate.objects.update_or_create(name=value[0], author=surveyProfile_object, date_created=datetime.date.today(), defaults={'name': value[0], 'author': surveyProfile_object, 'date_created': datetime.date.today()})
            if key.startswith("Section_"):
                if created:
                    new_section = Section.objects.create(section_name=value[0])
                    new_section.section_questionnaire.add(new_questionnaire)
                else:
                    pass
            if key.startswith("Question"):
                if created:
                    if "Section" in key:
                        new_question = Question.objects.create(text=value[0])
                        new_question.section.add(new_section)
                        new_question.save()
                    if "AnswerType" in key:
                        new_question.type_of_answer = value[0]
                        new_question.save()
                    if "AnswerDetail" in key:
                        new_question.detail_of_answer = value[0]
                        new_question.save()

                        if new_question.type_of_answer == "Range":
                            rangedAnswer_object = RangedAnswer.objects.get(id=new_question.detail_of_answer)
                            rangedAnswer_object.question.add(new_question)

                        if new_question.type_of_answer == "Text":
                            textAnswer_object = TextAnswer.objects.get(id=new_question.detail_of_answer)
                            textAnswer_object.question.add(new_question)

                    if "Order" in key:
                        try:
                            new_question.order = int(value[0])
                            new_question.save()
                        except:
                            new_question.order = question_order
                            question_order += 1
                    if "Weight" in key:
                        try:
                            new_question.weight = float(value[0])
                            new_question.save()
                        except:
                            new_question.weight = float(1)
                            new_question.save()

                    if "Description" in key:
                            new_question.description = value[0].capitalize()

                else:
                    pass

        context = {
            'ranged_answers': ranged_answers,
            'created': True,
            'questionnaire_id': new_questionnaire.id,
        }

    else:

        context = {
            'ranged_answers': ranged_answers,
            'created': False,
        }

    text_limits = TextAnswer.objects.all().order_by('length_limit')

    context.update({
        'text_limits': text_limits,
    })
    return render(request, 'create_questionnaire.html', context)


def ranged_answers(request):

    context = { 
        'ranged_answers': RangedAnswer.objects.all()
    }

    if request.method == "POST":
        print(request.POST)

        strings = json.loads(request.POST.get('strings', False))

        if strings:
            for key, value in strings.items():
                if not value:
                    strings[key] = key
        else:
            strings = {}
            minimum = int(request.POST.get('minimum'))
            maximum = int(request.POST.get('maximum'))
            for i in range(minimum, maximum+1):
                strings[i] = i


        new_rangedAnswer = RangedAnswer.objects.create(numeric_low=request.POST.get("minimum"), numeric_high=request.POST.get("maximum"), numeric_zero=request.POST.get("zero"), strings=strings)
        

    if request.method == "GET":
        pass


    return render(request, 'ranged_answers.html', context)


def delete_range(request):

    RangedAnswer.objects.get(id=request.POST.get('ranged_answer_id')).delete()

    data = {
        'deleted': True,
    }

    return JsonResponse(data)


@login_required
def edit_questionnaire(request, questionnaire_id):

    surveyProfile_object = SurveyProfile.objects.get(user=request.user)
    questionnaireTemplate_object = QuestionnaireTemplate.objects.get(id=questionnaire_id)
    sections = questionnaireTemplate_object.section_set.all()   
    ranged_answers = RangedAnswer.objects.all()
    text_limits = TextAnswer.objects.all()
    question_order = 1

    if request.method == "GET":
        pass

    if request.method == "POST":

        if request.POST.get('SubmitType') == "Smazat":
            questionnaireTemplate_object.delete()
            return redirect('edit_questionnaire_list')

        if request.POST.get('SubmitType') == "Duplikovat":
            new_questionnaireTemplate_object = QuestionnaireTemplate(name=questionnaireTemplate_object.name + ' (kopie)')
            new_questionnaireTemplate_object.save()
            new_questionnaireTemplate_object.section_set.add(*questionnaireTemplate_object.section_set.all())

            return redirect('edit_questionnaire', new_questionnaireTemplate_object.id)


        original_id = questionnaireTemplate_object.id
        questionnaireTemplate_object.delete()

        data = json.dumps(dict(request.POST))
        data = json.loads(data)

        for key, value in data.items():
            if key == "name":
                new_questionnaire = QuestionnaireTemplate.objects.create(name=value[0], author=surveyProfile_object, id=original_id)
            if key.startswith("Section_"):
                new_section = Section.objects.create(section_name=value[0])
                new_section.section_questionnaire.add(new_questionnaire)
            if key.startswith("Question"):
                if "Section" in key:
                    new_question = Question.objects.create(text=value[0])
                    new_question.section.add(new_section)
                    new_question.save()
                if "AnswerType" in key:
                    new_question.type_of_answer = value[0]
                    new_question.save()
                if "AnswerDetail" in key:
                    new_question.detail_of_answer = value[0]
                    new_question.save()

                    if new_question.type_of_answer == "Range":
                        rangedAnswer_object = RangedAnswer.objects.get(id=new_question.detail_of_answer)
                        rangedAnswer_object.question.add(new_question)

                    if new_question.type_of_answer == "Text":
                        textAnswer_object = TextAnswer.objects.get(id=new_question.detail_of_answer)
                        textAnswer_object.question.add(new_question)

                if "Order" in key:
                    try:
                        new_question.order = int(value[0])
                        new_question.save()
                    except:
                        new_question.order = question_order
                        question_order += 1
                if "Weight" in key:
                    try:
                        new_question.weight = float(value[0])
                        new_question.save()
                    except:
                        new_question.weight = float(1)
                        new_question.save()

                if "Description" in key:
                    if value[0] == 'true':
                        new_question.description = True
                    else:
                        new_question.description = value[0]


    text_limits = TextAnswer.objects.all()

    context = {
        'questionnaire': questionnaireTemplate_object,
        'sections': sections,
        'ranged_answers': ranged_answers,
        'text_limits': text_limits,
    }

    return render(request, 'edit_questionnaire.html', context)


def check_questionnaire(request):
    return None


def evaluate_questionnaire(request):
    return None


def distribute_questionnaire(request):

    questionnaire_templates = QuestionnaireTemplate.objects.all()
    employees = Employee.objects.all()

    context = {
        'questionnaires': questionnaire_templates,
        'employees': employees,
    }

    if request.method == "POST":
        questionnaireTemplate = QuestionnaireTemplate.objects.get(id=request.POST.get('questionnaire'))

        questionnaireTemplate.locked = True
        questionnaireTemplate.save()

        for employee in employees:
            if employee.evaluator and not employee.evaluator == 'bez hodnotitele':
                try:
                    interviewerObject = Employee.objects.get(name=employee.evaluator)
                except MultipleObjectsReturned:
                    interviewerObject = Employee.objects.get(name=employee.evaluator, is_supervisor=True)
                interviewees = Employee.objects.filter(evaluator=interviewerObject.name)

                for interviewee in interviewees:
                    if not Questionnaire.objects.filter(interviewee=interviewee, questionnaire=questionnaireTemplate).exists():
                        Questionnaire.objects.create(
                            interviewer = interviewerObject,
                            questionnaire = questionnaireTemplate,
                            interviewee = interviewee,
                            )

        return_message = "Dotazník byl úspěšně rozeslán"

        context.update({
            'return_message': return_message,
        })

        
    return render(request, 'distribute_questionnaire.html', context)


def my_questionnaires(request, questionnairetemplate_id):

    try:
        supervisor = Employee.objects.get(name=str(request.user.last_name + ' ' + request.user.first_name))
    except:
        supervisor = Employee.objects.filter(name=str(request.user.last_name + ' ' + request.user.first_name), is_supervisor=True).first()
    interviewees = Employee.objects.filter(evaluator=supervisor.name)
    questionnaireTemplate = QuestionnaireTemplate.objects.get(id=questionnairetemplate_id)

    total_questions = 0

    sections = questionnaireTemplate.section_set.all()

    for section in sections:
        questions = section.question.all()
        total_questions += questions.count()

    context = {
        'questionnaires': Questionnaire.objects.filter(interviewee__in=interviewees, questionnaire=questionnaireTemplate).order_by('interviewee'),
        'questionnaire_template': questionnaireTemplate,
        'total_questions': total_questions,
    }

    if request.method == "POST":
        pass

    return render(request, 'my_questionnaires.html', context)


def change_answer(request, questionnaire_id, question_id):

    questionnaireObject = Questionnaire.objects.get(id=questionnaire_id)
    questionObject = Question.objects.get(id=question_id)
    range_answer = questionObject.ranged_answer.first()

    context = {
        'question': questionObject,
        'ranged_answer': range_answer,
        'current_answer': questionnaireObject.answers[str(question_id)],
    }

    if request.method == "POST":
        for key, value in request.POST.dict().items():
            if str(key) in questionnaireObject.answers:
                questionnaireObject.answers[str(key)] = value
                questionnaireObject.save()

        context.update({
            'current_answer': questionnaireObject.answers[str(question_id)],
        })

    return render(request, 'change_answer.html', context)


@login_required
def my_questionnaires_list(request):
    print(request.user.last_name)
    try:
        supervisor = Employee.objects.get(name=str(request.user.last_name + ' ' + request.user.first_name))
    except:
        supervisor = Employee.objects.filter(name=str(request.user.last_name + ' ' + request.user.first_name), is_supervisor=True).first()
    interviewees = Employee.objects.filter(evaluator=supervisor.name)

    context = {
        'questionnaires': Questionnaire.objects.filter(interviewee__in=interviewees, active=True).distinct('questionnaire'),
        'active_questionnaires_count': Questionnaire.objects.filter(interviewee__in=interviewees, active=True, sent=False),
        'inactive_questionnaires': Questionnaire.objects.filter(interviewee__in=interviewees, active=False).distinct('questionnaire')
    }

    return render(request, 'my_questionnaires_list.html', context)


@login_required
def fill_in_questionnaire(request, questionnaire_id):

    lastname = request.user.last_name
    firstname = request.user.first_name

    try:
        interviewer = Employee.objects.get(name=lastname + ' ' + firstname)
    except:
        interviewer = Employee.objects.filter(name=lastname + ' ' + firstname, is_supervisor=True).first()

    questionnaire = Questionnaire.objects.get(id=questionnaire_id)
    questionnaireTemplate = questionnaire.questionnaire

    if request.method == "POST":
        questionnaire.answers = json.loads(request.POST.get('strings'))
        questionnaire.descriptions = json.loads(request.POST.get('descriptions'))
        questionnaire.sent = True
        questionnaire.save()

        if not questionnaire.interviewer:
            questionnaire.interviewer = interviewer
            questionnaire.save()

    context = {
        'questionnaire': Questionnaire.objects.filter(id=questionnaire_id, active=True).first(),
        'descriptions_from_db': list(Questionnaire.objects.filter(id=questionnaire_id).values_list('descriptions', flat=True))[0],
        'period_over': datetime.datetime.now() <= datetime.datetime(2023, 3, 8)
    }

    if request.method == "GET":
        return render(request, 'fill_in_questionnaire.html', context)
    
    if request.method == "POST":
        return redirect('my_questionnaires', questionnaireTemplate.id)


@login_required
def fill_in_questionnaire_devel(request, questionnaire_id):

    lastname = request.user.last_name
    firstname = request.user.first_name
    
    try:
        interviewer = Employee.objects.get(name=lastname + ' ' + firstname)
    except:
        interviewer = Employee.objects.filter(name=lastname + ' ' + firstname, is_supervisor=True).first()

    questionnaire = Questionnaire.objects.get(id=questionnaire_id)
    questionnaireTemplate = questionnaire.questionnaire

    if request.method == "POST":
        questionnaire.answers = json.loads( request.POST.get('strings'))
        questionnaire.descriptions = json.loads(request.POST.get('descriptions'))
        questionnaire.sent = True
        questionnaire.save()

        if not questionnaire.interviewer:
            questionnaire.interviewer = interviewer
            questionnaire.save()

    context = {
        'questionnaire': Questionnaire.objects.filter(id=questionnaire_id, active=True).first(),
        'descriptions_from_db': list(Questionnaire.objects.filter(id=questionnaire_id).values_list('descriptions', flat=True))[0],
        'period_over': datetime.datetime.now().date() < datetime.datetime(2021, 3, 1).date(),
    }

    if request.method == "GET":
        return render(request, 'fill_in_questionnaire_devel.html', context)
    
    if request.method == "POST":
        return redirect('my_questionnaires', questionnaireTemplate.id)


def edit_filled_in_questionnaire(request, questionnaire_id):
    
    filled_in_questionaire = Questionnaire.objects.get(id=questionnaire_id)

    context = {
        'filled_in_questionnaire': filled_in_questionaire,
    }

    return render(request, 'edit_filled_in_questionnaire.html', context)


def edit_interviewees(request):
    return None


def validate_survey_name(request):

    survey_name = request.POST.get('survey_name')
    data = {
        'is_taken': QuestionnaireTemplate.objects.filter(name=survey_name).exists()
    }

    return JsonResponse(data)


def querydict_to_dict(query_dict):
    data = {}
    for key in query_dict.keys():
        v = query_dict.getlist(key)
        if len(v) == 1:
            v = v[0]
        data[key] = v
    return data


def show_questionnaire(request, questionnaire_id):

    questionnaireTemplate_object = QuestionnaireTemplate.objects.get(id=questionnaire_id)

    context = {
        'questionnaire_template': questionnaireTemplate_object,
    }

    return render(request, 'show_questionnaire.html', context)


def edit_questionnaire_list(request):

    questionnaires = QuestionnaireTemplate.objects.all()

    context = {
        'questionnaires': questionnaires,
    }

    return render(request, 'edit_questionnaire_list.html', context)


@login_required
def questionnaire_results(request):

    if request.user.last_name in ['Kuštanová', 'Töröková', 'Vajda', 'Černá']:

        questionnaireTemplates = QuestionnaireTemplate.objects.all()

        context = {
            'questionnaire_templates': questionnaireTemplates,
        }

        return render(request, 'questionnaire_results.html', context)

    else:

        return redirect('my_questionnaires_list')


@login_required
def questionnaire_results_questionnaire(request, questionnaire_id):

    if request.user.last_name in ['Kuštanová', 'Töröková', 'Vajda', 'Černá']:

        questionnaireTemplate = QuestionnaireTemplate.objects.get(id=questionnaire_id)

        questionnaires = Questionnaire.objects.filter(questionnaire=questionnaireTemplate).order_by('interviewee__cost_center_number', 'interviewee__name')

        interviewers_id = Questionnaire.objects.values_list('interviewer_id').distinct()

        interviewers_names = list(Employee.objects.filter(id__in=interviewers_id).values_list('name', flat=True).order_by('name'))
        cost_centers = Employee.objects.filter(id__in=Questionnaire.objects.select_related('interviewee').values_list('interviewee_id')).values_list('cost_center_number', flat=True).distinct('cost_center_number')

        context = {
            'questionnaires': questionnaires,
            'questionnaire_template': questionnaireTemplate,
            'names': interviewers_names,
            'cost_centers': cost_centers,
        }

        return render(request, 'questionnare_results_questionnaire.html', context)

    else:

        return redirect('my_questionnaires_list')



def questionnaire_results_questionnaire_charts(request, questionnaire_id):

    questionnaireTemplate = QuestionnaireTemplate.objects.get(id=questionnaire_id)

    questionnaires = Questionnaire.objects.filter(questionnaire=questionnaireTemplate).order_by('interviewee__cost_center_number', 'interviewee__name')

    statistics = {}

    statistics['per_department'] = {}

    departments = list(questionnaires.values_list('interviewee__cost_center_number', flat=True).distinct('interviewee__cost_center_number'))
    
    for department in departments:
        statistics['per_department'][department] = {}
        statistics['per_department'][department]['sent'] = questionnaires.filter(interviewee__cost_center_number=department, sent=True).count()
        statistics['per_department'][department]['unsent'] = questionnaires.filter(interviewee__cost_center_number=department, sent=False).count()
        statistics['per_department'][department]['ratio'] = statistics['per_department'][department]['sent'] / (statistics['per_department'][department]['sent'] + statistics['per_department'][department]['unsent']) * 100

    context = {
        'questionnaires': questionnaires,
        'questionnaires_total_count': questionnaires.count(),
        'questionnaires_sent': questionnaires.filter(sent=True).count(),
        'questionnaire_template': questionnaireTemplate,
        'statistics': statistics,
    }

    return render(request, 'questionnare_results_questionnaire_chart.html', context)


def employee_upload(request):
    
    context = {}

    if request.method == "POST":

        list_of_personal_numbers_from_database = Employee.objects.values_list('personal_number', flat=True)
        actual_personal_numbers = []

        for f in request.FILES.getlist('file'):
            personal_numbers_list = []
            wb = load_workbook(f)
            ws = wb.active
            row_index = 2
            while row_index <= ws.max_row:
                values = {}
                values['personal_number'] = ws['A' + str(row_index)].value
                actual_personal_numbers.append(values['personal_number'])
                values['name'] = ws['B' + str(row_index)].value
                values['cost_center_number'] = ws['C' + str(row_index)].value
                values['cost_center_name'] = ws['D' + str(row_index)].value
                personal_numbers_list.append(ws['A' + str(row_index)].value)
                values['position'] = ws['E' + str(row_index)].value
                values['is_supervisor'] = ws['F' + str(row_index)].value
                if values['is_supervisor'] in ["false", "Ne", "ne"]:
                    values['is_supervisor'] = False
                else:
                    values['is_supervisor'] = True

                values['evaluator'] = ws['G' + str(row_index)].value

                employee, created = Employee.objects.update_or_create(personal_number = values['personal_number'], defaults=values)
                row_index += 1

        personal_numbers_to_delete = list(set(list_of_personal_numbers_from_database) - set(actual_personal_numbers))

        Employee.objects.filter(personal_number__in=personal_numbers_to_delete).delete()

    return render(request, 'employee_upload.html', context)


def update_evaluators(request=None, interviewee_id=None):

    if interviewee_id:
        active_questionnaires = Questionnaire.objects.filter(active=True, interviewee_id=interviewee_id)
    else:
        active_questionnaires = Questionnaire.objects.filter(active=True)

    for active_questionnaire in active_questionnaires:
        active_questionnaire_current_evaluator = Employee.objects.get(id=active_questionnaire.interviewer_id).name
        active_questionnaire_interviewee = Employee.objects.get(id=active_questionnaire.interviewee_id)
        active_questionnaire_interviewee_evaluator = active_questionnaire_interviewee.evaluator
        active_questionnaire_interviewee_evaluator_id = Employee.objects.get(name=active_questionnaire_interviewee_evaluator, is_supervisor=True).id

        if not active_questionnaire_current_evaluator == active_questionnaire_interviewee_evaluator:
            active_questionnaire.interviewer_id = active_questionnaire_interviewee_evaluator_id
            active_questionnaire.save()

    return None